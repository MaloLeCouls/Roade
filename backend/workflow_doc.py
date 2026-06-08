"""Human-readable documentation of a workflow, as an Excel workbook.

The goal is a file you can hand to someone who never opens the app: for every
output file, it spells out — in plain French, step by step — every treatment that
leads to it, in execution order. One sheet of synthesis, then one sheet per output
file tracing its lineage from the source files down to the export.

Public API:
    build_workbook(pid, wid) -> bytes   # the .xlsx, ready to stream
"""
from __future__ import annotations

import math
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import storage
import engine
from engine import _sanitize_filename, _sanitize_sheet_name


# --------------------------------------------------------------------------- #
# Vocabulary
# --------------------------------------------------------------------------- #
TYPE_LABEL = {
    "source": "Source", "sql": "SQL", "dedup": "Doublons", "validate": "Validation",
    "pivot": "Pivot", "clean": "Nettoyage", "calc": "Calcul", "filter": "Filtre",
    "cols": "Colonnes", "union": "Union", "export": "Export",
}
TYPE_DESC = {
    "source": "Lecture d'un fichier Excel/CSV",
    "sql": "Transformation : sélection, filtre, jointure, regroupement",
    "dedup": "Détection / séparation des doublons",
    "validate": "Classement des lignes selon des conditions (conformité / aiguillage)",
    "pivot": "Pivot / dépivot (tableau croisé)",
    "clean": "Nettoyage des valeurs",
    "calc": "Colonnes calculées",
    "filter": "Filtrage par appartenance à un autre tableau",
    "cols": "Réorganisation / suppression / renommage de colonnes",
    "union": "Empilement de plusieurs tableaux",
    "export": "Écriture d'un fichier de sortie",
}

_AGG_FR = {
    "SUM": "somme", "AVG": "moyenne", "MIN": "minimum", "MAX": "maximum",
    "COUNT": "nombre", "COUNT_DISTINCT": "nombre de valeurs distinctes",
    "MEDIAN": "médiane", "STDDEV": "écart-type", "FIRST": "première valeur",
    "LAST": "dernière valeur", "STRING_AGG": "concaténation",
}
_OP_FR = {
    "=": "vaut", "==": "vaut", "!=": "est différent de", "<>": "est différent de",
    ">": "est supérieur à", ">=": "est supérieur ou égal à",
    "<": "est inférieur à", "<=": "est inférieur ou égal à",
    "LIKE": "ressemble à", "NOT LIKE": "ne ressemble pas à",
    "ILIKE": "ressemble à (sans casse)", "IN": "fait partie de",
    "NOT IN": "ne fait pas partie de", "IS NULL": "est vide",
    "IS NOT NULL": "n'est pas vide", "BETWEEN": "est compris entre",
}
_GROUP_FN_FR = {
    "occurrence": "n° d'occurrence dans le groupe", "group_size": "taille du groupe",
    "count_distinct": "nb de valeurs distinctes de", "is_duplicate": "valeur répétée dans le groupe",
    "sum": "somme de", "avg": "moyenne de", "min": "minimum de", "max": "maximum de",
    "median": "médiane de", "share": "part dans le total du groupe de",
    "first": "première valeur de", "last": "dernière valeur de",
    "prev": "valeur précédente de", "next": "valeur suivante de", "concat": "liste des valeurs de",
}
_GROUP_CHECK_FR = {
    "unique": "toutes les valeurs sont différentes (unicité)",
    "constant": "toutes les valeurs sont identiques (constance)",
    "distinct_eq": "exactement {n} valeurs distinctes",
    "distinct_min": "au moins {n} valeurs distinctes",
    "distinct_max": "au plus {n} valeurs distinctes",
    "contains": "contient la valeur « {v} »",
    "not_contains": "ne contient pas la valeur « {v} »",
    "size_eq": "exactement {n} lignes", "size_min": "au moins {n} lignes",
    "size_max": "au plus {n} lignes",
}


def _q(v) -> str:
    """A value, French-quoted; empty values read as « (vide) »."""
    s = "" if v is None else str(v)
    return f"« {s} »" if s != "" else "« (vide) »"


def _col(name) -> str:
    return f"« {name} »" if name else "la valeur"


# --------------------------------------------------------------------------- #
# Validation conditions — readable prose, with OR-consolidation
# --------------------------------------------------------------------------- #
# Tests whose OR can be safely merged into a single "one of: a, b, c" phrase.
_MERGEABLE = {"equals", "contains", "starts_with", "ends_with"}
_VERB_ONE = {"equals": "vaut", "contains": "contient",
             "starts_with": "commence par", "ends_with": "finit par"}
_VERB_MANY = {"equals": "vaut l'une des valeurs", "contains": "contient l'un de",
              "starts_with": "commence par l'un de", "ends_with": "finit par l'un de"}


_RULE_OPPOSITE = {
    "contains": "not_contains", "not_contains": "contains",
    "equals": "not_equals", "not_equals": "equals",
    "is_empty": "not_empty", "not_empty": "is_empty",
}


def _rule_phrase(r: dict, default_col) -> str:
    """One non-consolidated rule, e.g. « Code » commence par « FR ». A negated rule
    folds into its positive twin when it has one (NON + « ne contient pas » reads as
    « contient ») so the documentation never shows a double negative."""
    test = r.get("test")
    neg = ""
    if r.get("negate"):
        opp = _RULE_OPPOSITE.get(test)
        if opp:
            test = opp                          # « ne contient pas » negated -> « contient »
        else:
            neg = "NON — "
    v = r.get("value", "")
    col = r.get("column") or default_col
    start = int(r.get("start") or 1)
    length = int(r.get("length") or 1)
    cls = {"letter": "une lettre", "upper": "une majuscule", "lower": "une minuscule",
           "digit": "un chiffre", "alnum": "alphanumérique"}.get(r.get("charclass"), r.get("charclass") or "?")
    body = {
        "starts_with": f"{_col(col)} commence par {_q(v)}",
        "ends_with": f"{_col(col)} finit par {_q(v)}",
        "contains": f"{_col(col)} contient {_q(v)}",
        "not_contains": f"{_col(col)} ne contient pas {_q(v)}",
        "equals": f"{_col(col)} vaut {_q(v)}",
        "not_equals": f"{_col(col)} est différent de {_q(v)}",
        "regex": f"{_col(col)} correspond au motif {v}",
        "regex_full": f"{_col(col)} correspond entièrement au motif {v}",
        "length_eq": f"{_col(col)} fait {v} caractères",
        "length_min": f"{_col(col)} fait au moins {v} caractères",
        "length_max": f"{_col(col)} fait au plus {v} caractères",
        "char_class": f"le {start}ᵉ caractère de {_col(col)} est {cls}",
        "char_equals": f"le {start}ᵉ caractère de {_col(col)} est {_q(v)}",
        "substr_equals": f"les caractères {start} à {start + length - 1} de {_col(col)} valent {_q(v)}",
        "substr_class": f"les caractères {start} à {start + length - 1} de {_col(col)} sont {cls}",
        "is_in": f"{_col(col)} fait partie de la liste : {v}",
        "is_empty": f"{_col(col)} est vide",
        "not_empty": f"{_col(col)} n'est pas vide",
        "is_numeric": f"{_col(col)} est numérique",
    }.get(test, f"{_col(col)} : {test}")
    return neg + body


def _describe_rules(cond: dict, default_col) -> list[str]:
    """A 'rules' condition (a DNF: OR of groups, each group an AND of rules) turned
    into prose. The common pattern — many single-rule OR groups on the same column
    and test (« contient X » OU « contient Y » OU …) — is consolidated into a single
    « contient l'un de : X, Y, … » line."""
    groups = cond.get("groups") or []
    if not groups:
        return ["(aucune règle — ne filtre rien)"]

    disjuncts = []                      # ordered list of phrase strings (the OR terms)
    buckets = {}                        # (col, test) -> index into disjuncts (placeholder)
    bucket_vals = {}                    # (col, test) -> [values]

    for g in groups:
        rules = g.get("rules") or []
        if not rules:
            continue
        if len(rules) == 1:
            r = rules[0]
            col = r.get("column") or cond.get("column") or default_col
            test = r.get("test")
            mergeable = (test in _MERGEABLE and not r.get("negate")
                         and r.get("value") not in (None, ""))
            if test == "is_in" and not r.get("negate"):    # a list test folds into equals
                key = (col, "equals")
                if key not in buckets:
                    buckets[key] = len(disjuncts); disjuncts.append(None); bucket_vals[key] = []
                bucket_vals[key].extend(p.strip() for p in str(r.get("value", "")).split(",") if p.strip())
                continue
            if mergeable:
                key = (col, test)
                if key not in buckets:
                    buckets[key] = len(disjuncts); disjuncts.append(None); bucket_vals[key] = []
                bucket_vals[key].append(r.get("value"))
                continue
            disjuncts.append(_rule_phrase(r, default_col))
        else:                            # an AND group: keep the conjunction explicit
            sub = " ET ".join(_rule_phrase(r, default_col) for r in rules)
            disjuncts.append(f"({sub})")

    # materialize the consolidated buckets back into their reserved slots
    for (col, test), pos in buckets.items():
        vals = bucket_vals[(col, test)]
        if len(vals) == 1:
            disjuncts[pos] = f"{_col(col)} {_VERB_ONE.get(test, test)} {_q(vals[0])}"
        else:
            lst = ", ".join(_q(v) for v in vals)
            disjuncts[pos] = f"{_col(col)} {_VERB_MANY.get(test, test)} : {lst}"

    disjuncts = [d for d in disjuncts if d]
    if len(disjuncts) == 1:
        return [disjuncts[0]]
    return ["L'UNE de ces conditions :"] + [f"   • {d}" for d in disjuncts]


def _describe_condition(cond: dict, default_col) -> list[str]:
    kind = cond.get("kind") or "rules"
    if kind == "mask":
        segs = cond.get("segments") or []
        if not segs:
            return ["doit être vide"]
        parts = [_seg_label(s) for s in segs]
        return [f"{_col(cond.get('column') or default_col)} suit le format : " + " + ".join(parts)]
    if kind == "group":
        keys = ", ".join(_q(c) for c in (cond.get("group_by") or [])) or "(toute la table)"
        tmpl = _GROUP_CHECK_FR.get(cond.get("check") or "unique", cond.get("check") or "?")
        check = tmpl.format(n=cond.get("n"), v=cond.get("value"))
        col = cond.get("column")
        coltxt = f" sur {_col(col)}" if col and "{v}" not in tmpl else ""
        return [f"par groupe de {keys} : {check}{coltxt}"]
    return _describe_rules(cond, default_col)


def _seg_label(seg: dict) -> str:
    t = seg.get("type") or "any"
    names = {"literal": None, "letter": "lettres", "upper": "MAJ", "lower": "min",
             "digit": "chiffres", "alnum": "alphanum.", "set": f"[{seg.get('value') or ''}]", "any": "tout"}
    if t == "literal":
        return _q(seg.get("value") or "")
    base = names.get(t, t)
    ln, lo, hi = seg.get("length"), seg.get("min"), seg.get("max")
    if ln not in (None, ""):
        return f"{base}×{int(ln)}"
    if lo not in (None, "") or hi not in (None, ""):
        return f"{base}×{lo or 0}–{hi if hi not in (None, '') else '∞'}"
    return base


# --------------------------------------------------------------------------- #
# Per-block descriptions  ->  list of lines (first line = headline)
# --------------------------------------------------------------------------- #
def _d_source(d, ins, a2l, wf):
    out = [f"Lit le fichier source {_q(d.get('file') or '(non choisi)')}."]
    if d.get("sheet"):
        out.append(f"Feuille : {_q(d.get('sheet'))}.")
    if d.get("header_row"):
        out.append(f"Ignore {int(d['header_row'])} ligne(s) avant les en-têtes de colonnes.")
    return out


def _sql_select(items) -> str:
    if not items:
        return "Toutes les colonnes"
    parts = []
    for it in items:
        kind = it.get("kind", "column")
        alias = it.get("alias")
        if kind == "expr":
            base = f"calcul : {it.get('expression', '').strip()}"
        elif kind == "agg":
            fn = _AGG_FR.get((it.get("func") or "").upper(), it.get("func"))
            base = f"{fn} de {_col(it.get('column'))}" if it.get("column") else fn
        else:
            base = _col(it.get("column"))
        if alias:
            base += f" → {_q(alias)}"
        parts.append(base)
    return ", ".join(parts)


def _sql_condition(c, a2l) -> str:
    op = (c.get("op") or "=").upper()
    if c.get("func"):
        fn = _AGG_FR.get(c["func"].upper(), c["func"])
        left = f"{fn} de {_col(c.get('column'))}"
    else:
        left = _col(c.get("column"))
    verb = _OP_FR.get(op, op.lower())
    if op in ("IS NULL", "IS NOT NULL"):
        return f"{left} {verb}"
    if op in ("IN", "NOT IN"):
        vals = ", ".join(_q(p.strip()) for p in str(c.get("value", "")).split(",") if p.strip())
        return f"{left} {verb} : {vals}"
    if op == "BETWEEN":
        return f"{left} {verb} {_q(c.get('value'))} et {_q(c.get('value2'))}"
    return f"{left} {verb} {_q(c.get('value'))}"


def _sql_chain(conds, a2l) -> str:
    out = ""
    for i, c in enumerate(conds):
        piece = _sql_condition(c, a2l)
        if i == 0:
            out = piece
        else:
            conn = "OU" if (c.get("connector") or "AND").upper() == "OR" else "ET"
            out += f" {conn} {piece}"
    return out


def _d_sql(d, ins, a2l, wf):
    if d.get("mode") == "raw":
        sql = (d.get("raw_sql") or "").strip()
        return ["Requête SQL personnalisée :"] + [f"   {ln}" for ln in sql.splitlines()]
    q = d.get("query") or {}
    out = [f"Colonnes en sortie : {_sql_select(q.get('select') or [])}."]
    if q.get("distinct"):
        out.append("Supprime les lignes en double (DISTINCT).")
    for j in q.get("joins") or []:
        jtype = (j.get("type") or "INNER").upper()
        tgt = a2l.get(j.get("input"), j.get("input"))
        if jtype == "CROSS":
            out.append(f"Jointure croisée (toutes les combinaisons) avec {_q(tgt)}.")
        else:
            conds = " et ".join(f"{_col(o.get('left_column'))} = {_col(o.get('right_column'))}"
                                for o in (j.get("on") or []))
            kind = {"INNER": "interne", "LEFT": "à gauche", "RIGHT": "à droite", "FULL": "complète"}.get(jtype, jtype)
            out.append(f"Jointure {kind} avec {_q(tgt)} (appariées quand {conds}).")
    if q.get("where"):
        out.append("Ne conserve que les lignes où : " + _sql_chain(q["where"], a2l) + ".")
    if q.get("group_by"):
        cols = ", ".join(_col(g.get("column")) for g in q["group_by"])
        out.append(f"Regroupe par {cols} (une ligne par combinaison).")
    if q.get("having"):
        out.append("Ne garde que les groupes où : " + _sql_chain(q["having"], a2l) + ".")
    if q.get("order_by"):
        parts = [f"{_col(o.get('column'))} ({'décroissant' if (o.get('dir') or 'ASC').upper() == 'DESC' else 'croissant'})"
                 for o in q["order_by"]]
        out.append("Trie par " + ", ".join(parts) + ".")
    if q.get("limit") not in (None, "", 0, "0"):
        out.append(f"Limite à {q['limit']} ligne(s).")
    return out


def _d_dedup(d, ins, a2l, wf):
    keys = d.get("key_columns") or []
    keytxt = ", ".join(_q(k) for k in keys) if keys else "la ligne entière"
    keep = "la dernière" if d.get("keep") == "last" else "la première"
    dups = {"all": "toutes les occurrences des groupes en double",
            "exemplar": "un exemplaire par groupe en double",
            "extra": "seulement les occurrences en trop"}.get(d.get("dups_mode", "all"))
    return [
        f"Repère les doublons sur : {keytxt}.",
        f"Conserve {keep} occurrence de chaque groupe.",
        "Trois sorties — Dédoublonné (un par groupe), "
        f"Doublons ({dups}), Uniques (présents une seule fois).",
    ]


def _d_validate(d, ins, a2l, wf):
    intent = d.get("intent")
    out = ["Contrôle de conformité." if intent == "control" else "Aiguille les lignes selon des conditions."]
    if d.get("target_column"):
        out.append(f"Colonne contrôlée par défaut : {_q(d['target_column'])}.")
    if d.get("case_sensitive"):
        out.append("Comparaisons sensibles à la casse (majuscules / minuscules distinguées).")
    conds = {c.get("id"): c for c in (d.get("conditions") or []) if c.get("id")}
    default_col = d.get("target_column")
    for c in d.get("conditions") or []:
        lines = _describe_condition(c, default_col)
        out.append(f"Condition {_q(c.get('name') or c.get('id'))} : {lines[0]}")
        out.extend(f"   {ln}" for ln in lines[1:])
    for o in d.get("outputs") or []:
        m = o.get("match") or {}
        cond = conds.get(m.get("conditionId"))
        cname = cond.get("name") if cond else "(aucune condition)"
        sense = "ne vérifient PAS" if m.get("negate") else "vérifient"
        out.append(f"→ Sortie {_q(o.get('label') or o.get('id'))} : les lignes qui {sense} la condition {_q(cname)}.")
    if d.get("routing", "first") == "first":
        out.append("Chaque ligne est dirigée vers la première sortie dont la condition est vraie.")
    if d.get("else_enabled", True):
        out.append(f"→ Sortie {_q(d.get('else_label') or 'Non classé')} : les lignes ne correspondant à aucune condition.")
    if d.get("add_flag"):
        out.append("Ajoute une colonne « valide » (oui / non).")
    if d.get("add_reason"):
        out.append("Ajoute une colonne « motif » expliquant les non-conformités.")
    return out


def _d_pivot(d, ins, a2l, wf):
    if d.get("mode") == "unpivot":
        vcs = ", ".join(_q(c) for c in (d.get("value_columns") or []))
        return [f"Dépivote les colonnes {vcs} en deux colonnes : "
                f"{_q(d.get('name_column') or 'variable')} (le nom) et {_q(d.get('value_column') or 'valeur')} (la valeur)."]
    idx = ", ".join(_q(c) for c in (d.get("index_columns") or []))
    agg = _AGG_FR.get((d.get("agg") or "SUM").upper(), d.get("agg"))
    return [f"Pivote (tableau croisé) : garde {idx} en lignes, "
            f"éclate {_col(d.get('pivot_column'))} en colonnes, "
            f"valeurs = {agg} de {_col(d.get('value_column'))}."]


_CLEAN_FR = {
    "trim": "supprime les espaces de début/fin", "upper": "met en MAJUSCULES",
    "lower": "met en minuscules", "capitalize": "met une Majuscule en tête",
    "replace": "remplace du texte", "fillna": "remplit les valeurs vides",
    "to_number": "convertit en nombre", "to_integer": "convertit en entier",
    "to_date": "convertit en date", "round": "arrondit", "abs": "valeur absolue",
}


def _d_clean(d, ins, a2l, wf):
    ops = [o for o in (d.get("operations") or []) if o.get("enabled") is not False]
    if not ops:
        return ["Aucune opération de nettoyage active."]
    out = [f"Applique {len(ops)} opération(s) de nettoyage, dans l'ordre :"]
    for i, o in enumerate(ops, 1):
        label = _CLEAN_FR.get(o.get("op"), o.get("op"))
        line = f"   {i}. {label} sur {_col(o.get('column'))}"
        if o.get("op") == "replace":
            line += f" — {_q(o.get('find'))} → {_q(o.get('replace'))}" + (" (regex)" if o.get("regex") else "")
        elif o.get("op") == "fillna":
            line += f" — par {_q(o.get('value'))}"
        elif o.get("op") == "to_date" and o.get("format"):
            line += f" — format {o.get('format')}"
        elif o.get("op") == "round":
            line += f" — {o.get('digits', 0)} décimale(s)"
        out.append(line)
    return out


def _calc_op_text(item) -> str:
    """Plain-French gloss of one calculated column, from its visual op model when
    present, otherwise the (already readable) formula string."""
    op = item.get("op") or {}
    kind = op.get("kind") or "formula"
    c = op.get("column")
    if kind == "extract":
        w = op.get("what") or "after_sep"
        what = {"after_sep": f"le texte après {_q(op.get('sep'))}",
                "before_sep": f"le texte avant {_q(op.get('sep'))}",
                "first_n": f"les {op.get('n', 1)} premiers caractères",
                "last_n": f"les {op.get('n', 1)} derniers caractères",
                "mid": f"{op.get('n', 5)} caractères à partir de la position {op.get('start', 1)}"}.get(w, w)
        return f"extrait {what} de {_col(c)}"
    if kind == "transform":
        t = {"upper": "en MAJUSCULES", "lower": "en minuscules", "trim": "sans espaces superflus",
             "replace": f"en remplaçant {_q(op.get('find'))} par {_q(op.get('replace'))}"}.get(op.get("transform"), op.get("transform"))
        return f"{_col(c)} {t}"
    if kind == "combine":
        segs = [(_q(p.get("value")) if p.get("type") == "text" else _col(p.get("value"))) for p in (op.get("parts") or [])]
        sep = f" séparées par {_q(op.get('sep'))}" if op.get("sep") else ""
        return "colle " + " + ".join(segs) + sep
    if kind == "condition":
        return f"si {_col(c)} {op.get('cmp', '=')} {_q(op.get('value'))} alors {_q(op.get('then'))} sinon {_q(op.get('else'))}"
    if kind == "math":
        rhs = _col(op.get("rhs")) if op.get("rhsIsCol") else op.get("rhs")
        r = f" arrondi à {op.get('round')} déc." if op.get("round") not in (None, "") else ""
        return f"{_col(c)} {op.get('mathOp', '+')} {rhs}{r}"
    return f"formule : {item.get('expr', '')}"


def _d_calc(d, ins, a2l, wf):
    out = []
    g = d.get("group") or {}
    gfuncs = [f for f in (g.get("functions") or []) if f.get("enabled") is not False]
    if gfuncs:
        keys = ", ".join(_q(c) for c in (g.get("partition_by") or [])) or "(toute la table)"
        out.append(f"Par groupe de {keys}, ajoute :")
        for f in gfuncs:
            fn = _GROUP_FN_FR.get(f.get("fn"), f.get("fn"))
            tgt = f" {_col(f.get('column'))}" if f.get("column") else ""
            out.append(f"   • {_q(f.get('name'))} = {fn}{tgt}")
    cols = [c for c in (d.get("columns") or []) if c.get("enabled") is not False and (c.get("name") or "").strip()]
    if cols:
        out.append(f"Ajoute {len(cols)} colonne(s) calculée(s) :")
        for c in cols:
            out.append(f"   • {_q(c.get('name'))} = {_calc_op_text(c)}")
    if not out:
        out.append("Aucune colonne calculée active.")
    return out


def _d_filter(d, ins, a2l, wf):
    ref = a2l.get("ref", "(référence non branchée)")
    keep = d.get("mode", "keep") != "exclude"
    verb = "Ne conserve que" if keep else "Retire"
    out = [f"{verb} les lignes dont {_col(d.get('column'))} figure dans "
           f"{_col(d.get('ref_column'))} du tableau de référence {_q(ref)}."]
    out.append("Type : semi-jointure (filtre par présence)." if keep
               else "Type : anti-jointure (filtre par absence).")
    if d.get("case_insensitive"):
        out.append("Comparaison insensible à la casse.")
    out.append("Aucune colonne de la référence n'est ajoutée — seules les lignes sont filtrées.")
    return out


def _d_cols(d, ins, a2l, wf):
    cols = d.get("columns") or []
    kept = [c for c in cols if c.get("keep") is not False]
    dropped = [c for c in cols if c.get("keep") is False]

    def lbl(c):
        rn = (c.get("rename") or "").strip()
        return f"{_q(c['name'])} → {_q(rn)}" if rn and rn != c["name"] else _q(c["name"])

    out = ["Réorganise / sélectionne les colonnes."]
    if kept:
        out.append("Conserve, dans l'ordre : " + ", ".join(lbl(c) for c in kept) + ".")
    if dropped:
        out.append("Supprime : " + ", ".join(_q(c["name"]) for c in dropped) + ".")
    return out


def _d_union(d, ins, a2l, wf):
    src = ", ".join(_q(s) for s in ins) if ins else "(aucune entrée)"
    align = "par position" if d.get("by_name") is False else "par nom de colonne"
    out = [f"Empile (concatène) les lignes de : {src}.", f"Alignement des colonnes : {align}."]
    if d.get("distinct"):
        out.append("Supprime les lignes en double après empilement.")
    return out


def _output_target(d, wf) -> str:
    """Short, unambiguous name of what an Export block writes: a standalone file,
    or a named sheet inside the workflow's workbook."""
    fn = _sanitize_filename(d.get("filename"), "resultat")
    if d.get("to_workbook"):
        return f"{_sanitize_sheet_name(fn)} (feuille du classeur « {wf}.xlsx »)"
    ext = "csv" if (d.get("format") or "xlsx").lower() == "csv" else "xlsx"
    return f"{fn}.{ext}"


def _d_export(d, ins, a2l, wf):
    fn = _sanitize_filename(d.get("filename"), "resultat")
    if d.get("to_workbook"):
        target = f"feuille {_q(_sanitize_sheet_name(fn))} du classeur {_q(wf + '.xlsx')}"
    else:
        ext = "csv" if (d.get("format") or "xlsx").lower() == "csv" else "xlsx"
        target = _q(f"{fn}.{ext}")
    out = [f"Écrit le résultat dans {target}."]
    if d.get("enabled") is False:
        out.append("⚠ Export désactivé : ce fichier n'est pas généré lors d'une exécution normale.")
    out.append("Un export sans données (0 ligne) ne crée aucun fichier.")
    return out


_DESCRIBERS = {
    "source": _d_source, "sql": _d_sql, "dedup": _d_dedup, "validate": _d_validate,
    "pivot": _d_pivot, "clean": _d_clean, "calc": _d_calc, "filter": _d_filter,
    "cols": _d_cols, "union": _d_union, "export": _d_export,
}


def _describe_block(node, ins, a2l, wf) -> list[str]:
    fn = _DESCRIBERS.get(node["type"])
    if not fn:
        return [f"Bloc « {node['type']} »."]
    try:
        return fn(node.get("data") or {}, ins, a2l, wf) or ["—"]
    except Exception as e:  # noqa: BLE001 - never let one odd block break the whole doc
        return [f"(description indisponible : {e})"]


# --------------------------------------------------------------------------- #
# Workbook assembly
# --------------------------------------------------------------------------- #
_TITLE = Font(bold=True, size=14, color="1F2A44")
_META = Font(size=10, italic=True, color="6B7280")
_SECTION = Font(bold=True, size=11, color="1F2A44")
_SECTION_FILL = PatternFill("solid", fgColor="EEF1F6")
_HEAD = Font(bold=True, size=10, color="1F2A44")
_HEAD_FILL = PatternFill("solid", fgColor="E2E7F0")
_BODY = Font(size=10, color="222933")
_BODY_STRONG = Font(size=10, bold=True, color="222933")
_TOPWRAP = Alignment(vertical="top", wrap_text=True)
_TOP = Alignment(vertical="top")
_TOPCENTER = Alignment(vertical="top", horizontal="center")
_thin = Side(style="thin", color="D7DCE5")
_ROW_BORDER = Border(bottom=_thin)
_HEAD_BORDER = Border(bottom=Side(style="medium", color="9AA6BC"))

_STEP_COLS = [("Étape", 7), ("Bloc", 24), ("Type", 13), ("Entrées", 24), ("Traitement", 90)]
_DESC_WRAP = 92                                    # ~chars before the Traitement cell wraps


def _autoheight(ws, row, text):
    lines = sum(max(1, math.ceil((len(ln) + 1) / _DESC_WRAP)) for ln in str(text).split("\n"))
    ws.row_dimensions[row].height = min(409, max(16, lines * 14.5))


def _unique_sheet_name(base: str, used: set) -> str:
    base = _sanitize_sheet_name(base)
    name, k = base, 2
    while name.lower() in used:
        sfx = f" ({k})"
        name = base[:31 - len(sfx)] + sfx
        k += 1
    used.add(name.lower())
    return name


def _steps_table(ws, start_row, step_ids, nodes, incoming, label_of, wf_name):
    """Write the Étape / Bloc / Type / Entrées / Traitement table and return the
    next free row."""
    for j, (title, width) in enumerate(_STEP_COLS, 1):
        ws.column_dimensions[get_column_letter(j)].width = width
        c = ws.cell(start_row, j, title)
        c.font = _HEAD; c.fill = _HEAD_FILL; c.alignment = _TOP; c.border = _HEAD_BORDER
    ws.freeze_panes = ws.cell(start_row + 1, 1)

    r = start_row + 1
    for i, nid in enumerate(step_ids, 1):
        node = nodes[nid]
        ins = [label_of(s) for _, s in incoming[nid]]
        a2l = {alias: label_of(s) for alias, s in incoming[nid]}
        text = "\n".join(_describe_block(node, ins, a2l, wf_name))
        ws.cell(r, 1, i).font = _BODY
        ws.cell(r, 2, label_of(nid)).font = _BODY_STRONG
        ws.cell(r, 3, TYPE_LABEL.get(node["type"], node["type"])).font = _BODY
        ws.cell(r, 4, ", ".join(ins) if ins else "—").font = _BODY
        ws.cell(r, 5, text).font = _BODY
        for j in range(1, 6):
            cell = ws.cell(r, j)
            cell.border = _ROW_BORDER
            cell.alignment = _TOPCENTER if j == 1 else (_TOPWRAP if j in (4, 5) else _TOP)
        _autoheight(ws, r, text)
        r += 1
    return r


def build_workbook(pid: str, wid: str) -> bytes:
    wf = storage.get_workflow(pid, wid)
    if not wf:
        raise ValueError("Workflow introuvable")
    name = (wf.get("name") or "Workflow").strip() or "Workflow"

    nodes = {n["id"]: n for n in wf.get("nodes", []) if n.get("type") not in ("frame", "group")}
    edges = [e for e in wf.get("edges", []) if e.get("source") in nodes and e.get("target") in nodes]
    order = engine._topo_order(nodes, edges)

    incoming = {nid: [] for nid in nodes}             # node -> [(handle, source_id)]
    for e in edges:
        incoming[e["target"]].append((e.get("targetHandle") or "in", e.get("source")))
    for nid in incoming:
        incoming[nid].sort(key=lambda x: x[0])

    def label_of(nid):
        n = nodes[nid]
        return (n["data"].get("label") or TYPE_LABEL.get(n["type"], n["type"]))

    wb = Workbook()
    synth = wb.active
    synth.title = "Synthèse"

    export_ids = [nid for nid in order if nodes[nid]["type"] == "export"]
    used_names = {"synthèse"}
    tab_of = {}

    if export_ids:
        for eid in export_ids:
            d = nodes[eid].get("data") or {}
            base = _sanitize_sheet_name(_sanitize_filename(d.get("filename"), "resultat"))
            sn = _unique_sheet_name(base, used_names)
            tab_of[eid] = sn
            ws = wb.create_sheet(sn)
            anc = engine._ancestors(eid, edges) | {eid}
            steps = [nid for nid in order if nid in anc]

            ws.cell(1, 1, f"Fichier de sortie : {_output_target(d, name)}").font = _TITLE
            sub = f"Bloc Export « {label_of(eid)} » — {_describe_block(nodes[eid], [], {}, name)[0]}"
            ws.cell(2, 1, sub).font = _META
            ws.cell(3, 1, f"Produit par {len(steps)} étape(s) de traitement, dans l'ordre d'exécution ci-dessous :").font = _BODY
            _steps_table(ws, 5, steps, nodes, incoming, label_of, name)
    else:
        ws = wb.create_sheet(_unique_sheet_name("Traitements", used_names))
        ws.cell(1, 1, "Traitements du workflow").font = _TITLE
        ws.cell(2, 1, "Ce workflow n'a pas de bloc Export ; voici tous ses traitements, dans l'ordre.").font = _META
        _steps_table(ws, 4, order, nodes, incoming, label_of, name)

    _build_synthesis(synth, name, nodes, order, export_ids, tab_of, label_of)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_synthesis(ws, name, nodes, order, export_ids, tab_of, label_of):
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 60

    def section(row, text):
        c = ws.cell(row, 1, text)
        c.font = _SECTION
        for col in (1, 2, 3):
            ws.cell(row, col).fill = _SECTION_FILL
        return row + 1

    ws.cell(1, 1, f"Documentation du workflow : {name}").font = _TITLE
    ws.cell(2, 1, f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — document explicatif, "
                  "lecture sans l'application.").font = _META
    r = 4

    r = section(r, "Comment lire ce document")
    for line in (
        "• Une feuille par fichier de sortie (onglets en bas) détaille, étape par étape, tous les",
        "   traitements qui mènent à ce fichier — des fichiers sources jusqu'à l'écriture finale.",
        "• Les étapes sont listées dans l'ordre d'exécution ; la colonne « Traitement » explique chaque bloc.",
    ):
        ws.cell(r, 1, line).font = _BODY
        ws.cell(r, 1).alignment = Alignment(horizontal="left")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
    r += 1

    # Source files actually read
    src_files = []
    for nid in order:
        if nodes[nid]["type"] == "source":
            f = (nodes[nid].get("data") or {}).get("file")
            if f and f not in src_files:
                src_files.append(f)
    r = section(r, "Fichiers sources utilisés")
    if src_files:
        for f in src_files:
            ws.cell(r, 1, "•").font = _BODY
            ws.cell(r, 2, f).font = _BODY
            r += 1
    else:
        ws.cell(r, 2, "(aucun fichier source)").font = _META
        r += 1
    r += 1

    # Output files produced
    r = section(r, "Fichiers produits")
    head = ("Fichier de sortie", "Détaillé dans l'onglet")
    ws.cell(r, 2, head[0]).font = _HEAD
    ws.cell(r, 3, head[1]).font = _HEAD
    r += 1
    if export_ids:
        for eid in export_ids:
            d = nodes[eid].get("data") or {}
            target = _output_target(d, name)
            disabled = (d.get("enabled") is False)
            ws.cell(r, 1, "•").font = _BODY
            cell = ws.cell(r, 2, target + ("   (désactivé)" if disabled else ""))
            cell.font = _BODY
            ws.cell(r, 3, tab_of.get(eid, "")).font = _BODY
            r += 1
    else:
        ws.cell(r, 2, "(aucun bloc Export — voir l'onglet « Traitements »)").font = _META
        r += 1
    r += 1

    # Legend of block types actually used
    used_types = []
    for nid in order:
        t = nodes[nid]["type"]
        if t not in used_types:
            used_types.append(t)
    r = section(r, "Types de blocs utilisés")
    for t in used_types:
        ws.cell(r, 1, "•").font = _BODY
        ws.cell(r, 2, TYPE_LABEL.get(t, t)).font = _BODY_STRONG
        ws.cell(r, 3, TYPE_DESC.get(t, "")).font = _BODY
        r += 1
