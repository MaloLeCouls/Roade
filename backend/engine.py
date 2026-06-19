"""Workflow execution engine.

A workflow is a DAG of blocks (nodes) linked by edges. Each block reads the
Parquet output(s) of its upstream block(s), computes its own result and
materializes it to Parquet (one file per output *handle*) plus a cached
``.meta.json`` (columns / row_count / sample / stats).

Public API (consumed by main.py and the smoke tests in backend/_test_*.py):
    iter_run_workflow(pid, wid, only_node=None) -> generator of SSE events
    run_workflow(pid, wid, only_node=None)      -> {"ran": [...], "results": {...}}
    preview_node(pid, wid, node_id, handle="out", ...) -> dict
    column_profile(pid, wid, node_id, column, handle="out") -> dict
    peek_source(pid, file, sheet, header_row)   -> {"sheets", "columns"}
    validate_test(config, samples)              -> {"ok", "results"|"error"}
"""

from __future__ import annotations

import hashlib
import json
import math
import queue
import re
import threading

import charset_normalizer
import duckdb
import numpy as np
import pandas as pd

import query_builder
import storage
from formula import FormulaError, compile_formula
from query_builder import q_ident

# Output handles per block type (first one is the "primary" output).
OUTPUTS = {
    "source": ["out"],
    "sql": ["out"],
    "pivot": ["out"],
    "clean": ["out"],
    "calc": ["out"],
    "union": ["out"],
    "export": [],
    "filter": ["out"],
    "cols": ["out"],
    "report": ["out"],
    "dedup": ["kept", "dups", "uniques"],
    "validate": ["valid", "invalid"],
}


def _output_handles(node) -> list[str]:
    """Output handle ids for a node — dynamic for a validate block in 'route'
    mode (one handle per user-defined output, plus an optional 'else')."""
    if node["type"] == "validate" and node["data"].get("mode") == "route":
        ids = [o["id"] for o in (node["data"].get("outputs") or []) if o.get("id")]
        if node["data"].get("else_enabled", True):
            ids.append("else")
        return ids or ["else"]
    return OUTPUTS.get(node["type"], ["out"])


def prune_orphan_edges(wf: dict) -> list[dict]:
    """Drop any edge whose `sourceHandle` is no longer exposed by the source
    node (typical after a Validation route had an output removed/renamed). The
    UI's "Sortie X" pastille is gone but the edge survived in the JSON, where
    it would silently re-ingest a stale parquet at the next run. Returns the
    list of removed edges (for logging/diagnostics)."""
    nodes = {n["id"]: n for n in wf.get("nodes", []) or []}
    kept, removed = [], []
    for e in wf.get("edges", []) or []:
        src = nodes.get(e.get("source"))
        sh = e.get("sourceHandle") or "out"
        # Unknown source node: leave the edge alone — node-deletion paths handle
        # that elsewhere and a missing node is a different kind of breakage.
        if src is None or sh in set(_output_handles(src)):
            kept.append(e)
        else:
            removed.append(e)
    wf["edges"] = kept
    return removed


_NUMERIC_TYPES = (
    "TINYINT",
    "SMALLINT",
    "INTEGER",
    "BIGINT",
    "HUGEINT",
    "UTINYINT",
    "USMALLINT",
    "UINTEGER",
    "UBIGINT",
    "UHUGEINT",
    "FLOAT",
    "REAL",
    "DOUBLE",
    "DECIMAL",
)


# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #
def _lit(path) -> str:
    """DuckDB string literal for a filesystem path (forward slashes on Windows)."""
    return "'" + str(path).replace("\\", "/").replace("'", "''") + "'"


def _is_null(v) -> bool:
    return v is None or (isinstance(v, float) and math.isnan(v)) or v is pd.NaT


def _json_safe(v):
    if _is_null(v):
        return None
    if isinstance(v, np.integer):
        return int(v)
    if isinstance(v, np.floating):
        f = float(v)
        return None if math.isnan(f) else f
    if isinstance(v, np.bool_):
        return bool(v)
    if isinstance(v, (pd.Timestamp,)):
        return v.isoformat(sep=" ")
    if isinstance(v, bytes):
        return v.decode("utf-8", "replace")
    return v


def _df_records(df: pd.DataFrame) -> list[dict]:
    return [{k: _json_safe(val) for k, val in rec.items()} for rec in df.to_dict("records")]


def _describe(con, rel: str) -> list[dict]:
    rows = con.execute(f"DESCRIBE SELECT * FROM {rel}").fetchall()
    return [{"name": r[0], "type": r[1]} for r in rows]


def _is_numeric_type(t: str) -> bool:
    t = (t or "").upper()
    return any(t.startswith(p) for p in _NUMERIC_TYPES)


# --------------------------------------------------------------------------- #
# Source reading
# --------------------------------------------------------------------------- #
_EXCEL_EXT = (".xlsx", ".xls", ".xlsm")

# Sniffing FR/EN numbers : « 1 234,56 » / « 12,5 » vs « 1,234.56 » / « 12.5 ».
# Espaces classique + insécable ( ) + fine insécable ( ) — tous courants
# côté Excel FR.
_FR_NUM_RE = re.compile(r"^-?\d+(?:[\s  ]\d{3})*,\d+$")
_EN_NUM_RE = re.compile(r"^-?\d+(?:,\d{3})*\.\d+$")


def _sniff_encoding(path) -> str:
    """Detect text-file encoding from the first ~16 KB. Falls back to utf-8.

    Restreint le sniffer aux encodings plausibles dans un contexte Roade (FR-first
    + exports Excel Windows). Sur des fichiers très courts, charset-normalizer
    peut autrement classer une chaîne accentuée comme Big5 / Shift_JIS et faire
    pourrir le décodage — ici on évite ça en limitant le champ.
    """
    with path.open("rb") as f:
        raw = f.read(1 << 14)
    if not raw:
        return "utf-8"
    plausible = ["utf-8", "utf-16", "cp1252", "iso-8859-1", "iso-8859-15"]
    best = charset_normalizer.from_bytes(raw, cp_isolation=plausible).best()
    return (best.encoding if best else None) or "utf-8"


# Token isolant un nombre décimal dans une ligne brute ; les lookarounds excluent
# les milliers (`1234,56` n'est pas pris pour FR si entouré d'autres chiffres
# côté virgule). On regarde le texte directement, AVANT que pandas choisisse un
# séparateur — sinon un `12,5` se ferait avaler comme deux colonnes EN.
_FR_TOKEN_RE = re.compile(r"(?<!\d)-?\d{1,3}(?:[\s  ]\d{3})*,\d+(?!\d)")
_EN_TOKEN_RE = re.compile(r"(?<!\d)-?\d{1,3}(?:,\d{3})*\.\d+(?!\d)")


def _sniff_decimal(text: str) -> str | None:
    """Look at ~50 data lines of raw text. Return ',' if FR-style decimals
    dominate, '.' if EN-style dominate, None if ambiguous."""
    fr = en = 0
    for line in text.splitlines()[1:51]:  # skip header
        fr += len(_FR_TOKEN_RE.findall(line))
        en += len(_EN_TOKEN_RE.findall(line))
    # Seuil bas (2 valeurs typées suffisent à signer le format), mais on exige
    # une domination nette du camp gagnant pour ne pas trancher sur du bruit.
    if fr >= 2 and fr > en * 2:
        return ","
    if en >= 2 and en > fr * 2:
        return "."
    return None


def _resolve_read_options(path, data: dict) -> tuple[dict, dict]:
    """Compute the effective `pd.read_csv` kwargs for this source + a `detected`
    dict reporting what was sniffed (for the Source aperçu, todo A.3).

    `data` keys honored:
      - encoding  : 'auto' (default) or a Python codec name (e.g. 'cp1252')
      - decimal   : 'auto' (default), ',' or '.'
      - thousands : optional, any string (no sniffing — too ambiguous)
    """
    ext = path.suffix.lower()
    detected: dict = {}
    opts: dict = {}
    if ext in _EXCEL_EXT or ext == ".parquet":
        return opts, detected  # binary formats : nothing to sniff/choose here

    user_enc = (data.get("encoding") or "auto").strip().lower()
    if user_enc in ("", "auto"):
        enc = _sniff_encoding(path)
        detected["encoding"] = enc
    else:
        enc = user_enc
    opts["encoding"] = enc

    user_dec = (data.get("decimal") or "auto").strip().lower()
    if user_dec in ("", "auto"):
        try:
            with path.open("r", encoding=opts["encoding"], errors="replace") as f:
                head = "".join(next(f, "") for _ in range(51))
            dec = _sniff_decimal(head)
            if dec:
                detected["decimal"] = dec
        except Exception:  # noqa: BLE001 — sniffing is best-effort
            dec = None
    elif user_dec in (",", "."):
        dec = user_dec
    else:
        dec = None
    if dec:
        opts["decimal"] = dec

    thou = data.get("thousands")
    if thou:
        opts["thousands"] = str(thou)

    return opts, detected


# A.5 — cast explicite par colonne au plus tôt (à la lecture de la source),
# pour ne pas obliger l'utilisateur à insérer un bloc Nettoyage juste pour
# typer ses colonnes. A.6 — y inclut le parsing de dates (format explicite ou
# `dayfirst=True` par défaut, contexte FR).
_BOOL_MAP = {
    "true": True,
    "vrai": True,
    "oui": True,
    "yes": True,
    "1": True,
    "false": False,
    "faux": False,
    "non": False,
    "no": False,
    "0": False,
}


def _apply_casts(df: pd.DataFrame, casts, decimal: str, thousands) -> tuple[pd.DataFrame, list]:
    """Apply user-defined per-column casts. Returns the new df + a report
    `[{column, type, failed}]` exposé dans l'aperçu pour signaler les ratés
    (cf. todo A.5)."""
    report = []
    if not casts:
        return df, report
    for c in casts:
        col = c.get("column")
        kind = (c.get("type") or "").lower()
        if not col or col not in df.columns or kind in ("", "auto"):
            continue
        before_na = int(df[col].isna().sum())
        if kind == "number":
            s = df[col].astype(str)
            if thousands:
                s = s.str.replace(thousands, "", regex=False)
            # Si la colonne avait déjà été reparsée à `,` par pandas (decimal=',')
            # cette substitution est sans effet ; sinon on uniformise.
            if decimal == ",":
                s = s.str.replace(",", ".", regex=False)
            new = pd.to_numeric(s, errors="coerce")
        elif kind == "date":
            fmt = c.get("format") or None
            try:
                new = pd.to_datetime(df[col], format=fmt, errors="coerce", dayfirst=True)
            except Exception:  # noqa: BLE001
                new = pd.to_datetime(df[col], errors="coerce", dayfirst=True)
        elif kind == "text":
            # `astype(str)` convertit NaN en "nan" — on garde les vides.
            new = df[col].where(df[col].notna(), None).astype(object)
            mask = new.notna()
            new = new.where(~mask, new[mask].astype(str))
        elif kind == "boolean":
            new = df[col].astype(str).str.strip().str.lower().map(_BOOL_MAP)
        else:
            continue
        # Pour `text` on ne compte pas les ratés (il n'y en a pas).
        if kind == "text":
            failed = 0
        else:
            after_na = int(new.isna().sum())
            failed = max(0, after_na - before_na)
        df[col] = new
        report.append({"column": col, "type": kind, "failed": failed})
    return df, report


def _normalize_mixed_object_columns(df: pd.DataFrame) -> list[dict]:
    """openpyxl renvoie le type Python natif de chaque cellule — un fichier
    Excel humain mélange souvent texte et nombres dans la même colonne (ex. un
    libellé qui contient parfois juste `7`). pyarrow refuse alors de sérialiser
    la colonne object hétérogène (« Expected bytes, got a 'int' object ») et
    `to_parquet` plante avant tout traitement aval.

    Solution la moins traîtresse : on convertit en texte les colonnes object
    qui mélangent plusieurs types Python (en gardant les `None`/NaN tels
    quels). On ne touche pas aux colonnes monotypées (toutes-strings, toutes-
    dates…). Le rapport retourné permet de prévenir l'utilisateur (les
    nombres orphelins sont devenus du texte — un cast explicite après ce bloc
    est encore possible).
    """
    coerced: list[dict] = []
    for col in df.columns:
        if df[col].dtype != object:
            continue
        types = set()
        counts: dict[str, int] = {}
        for v in df[col]:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            # On range int/float/bool sous l'étiquette « nombre » : c'est la
            # confusion la plus fréquente (texte mêlé à des codes numériques),
            # et donner un compte « 138 textes / 7 nombres » est plus parlant
            # qu'« int / float / str ».
            if isinstance(v, bool):
                label = "booléen"
            elif isinstance(v, (int, float)):
                label = "nombre"
            elif isinstance(v, str):
                label = "texte"
            elif isinstance(v, bytes):
                label = "binaire"
            else:
                label = type(v).__name__
            types.add(label)
            counts[label] = counts.get(label, 0) + 1
        if len(types) <= 1:
            continue
        # Coercion en texte, NaN/None préservés (pas de "None"/"nan" littéraux).
        df[col] = df[col].map(
            lambda v: None if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)
        )
        coerced.append({"column": col, "counts": counts})
    return coerced


def _source_fingerprint(pid, node) -> str | None:
    """Identity of a source's inputs: file + sheet + header + read options +
    size + mtime. Changes whenever the underlying file or its read options
    change, so a cached output can be safely reused only while this is identical
    (cf. todo A.9)."""
    d = node["data"]
    f = d.get("file")
    if not f:
        return None
    p = storage.files_dir(pid) / f
    if not p.exists():
        return None
    st = p.stat()
    opts_key = "|".join(str(d.get(k) or "") for k in ("encoding", "decimal", "thousands"))
    return (
        f"{f}|{d.get('sheet') or ''}|{int(d.get('header_row') or 0)}|"
        f"{opts_key}|{st.st_size}|{st.st_mtime_ns}"
    )


# --------------------------------------------------------------------------- #
# Change tracking — a node's "compute signature"
# --------------------------------------------------------------------------- #
_SIG_DROP_KEYS = {"test_samples"}  # UI-only, dropped at any depth
_SIG_DROP_TOP = {"label", "description", "locked", "cache"}  # cosmetic / handled apart


def _clean_for_sig(obj):
    if isinstance(obj, dict):
        return {k: _clean_for_sig(v) for k, v in obj.items() if k not in _SIG_DROP_KEYS}
    if isinstance(obj, list):
        return [_clean_for_sig(x) for x in obj]
    return obj


def _node_signature(pid, node, upstream_sigs) -> str:
    """Merkle-style signature of a node's *computed output*: its type, the config
    that affects computation, the signatures of its upstream inputs and — for a
    source — the file fingerprint. Same signature ⇒ same output, so a stored
    result can be reused safely."""
    d = {k: v for k, v in (node.get("data") or {}).items() if k not in _SIG_DROP_TOP}
    payload = json.dumps(_clean_for_sig(d), sort_keys=True, ensure_ascii=False, default=str)
    h = hashlib.sha256()
    h.update(node["type"].encode("utf-8"))
    h.update(b"\x00")
    h.update(payload.encode("utf-8"))
    if node["type"] == "source":
        h.update(b"\x00")
        h.update((_source_fingerprint(pid, node) or "").encode("utf-8"))
    for s in sorted(upstream_sigs):
        h.update(b"\x00")
        h.update(s.encode("utf-8"))
    return h.hexdigest()


def _signatures_for(pid, wid, nodes, edges, order) -> dict:
    """Current signature of every node, in topo order. A locked node that still
    has a materialized result contributes its *stored* (frozen) signature, so
    downstream blocks stay cached for as long as it stays locked."""
    sigs = {}
    for nid in order:
        node = nodes[nid]
        ups = []
        for e in edges:
            if e.get("target") == nid:
                src = e.get("source")
                ups.append(
                    f"{e.get('targetHandle') or 'in'}<-{src}:"
                    f"{e.get('sourceHandle') or 'out'}#{sigs.get(src, '')}"
                )
        sig = _node_signature(pid, node, ups)
        if node["data"].get("locked"):
            hs = _output_handles(node)
            h0 = hs[0] if hs else "out"
            m = storage.read_node_meta(pid, wid, nid, h0) or {}
            if m.get("node_signature") and storage.node_parquet(pid, wid, nid, h0).exists():
                sig = m["node_signature"]
        sigs[nid] = sig
    return sigs


def _count_source_data_rows(path, sheet, header_row: int, data: dict | None = None) -> int:
    """Nombre de lignes de données attendues (hors lignes d'en-tête ignorées)."""
    ext = path.suffix.lower()
    hr = int(header_row or 0)
    if ext in _EXCEL_EXT:
        try:
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                names = wb.sheetnames
                target = sheet if (sheet and sheet in names) else names[0]
                ws = wb[target]
                max_row = ws.max_row or 0
                # skiprows=hr : on saute hr lignes, la suivante est l'en-tête pandas.
                return max(0, max_row - hr - 1)
            finally:
                wb.close()
        except Exception:  # noqa: BLE001
            return 0
    if ext == ".parquet":
        try:
            import pyarrow.parquet as pq

            return int(pq.read_metadata(path).num_rows)
        except Exception:  # noqa: BLE001
            return 0
    try:
        opts, _ = _resolve_read_options(path, data or {})
        enc = opts.get("encoding") or "utf-8"
        lines = 0
        with path.open("r", encoding=enc, errors="replace") as f:
            for _ in f:
                lines += 1
        return max(0, lines - hr - 1)
    except Exception:  # noqa: BLE001
        return 0


def _read_csv_with_progress(path, hr, opts, ext, total, on_progress):
    kwargs = dict(skiprows=hr, **opts)
    if ext in (".tsv", ".txt"):
        kwargs["sep"] = "\t"
    else:
        kwargs["sep"] = None
        kwargs["engine"] = "python"
    chunk_size = max(1000, min(10000, total // 50 or 1000))
    chunks, rows_read = [], 0
    for chunk in pd.read_csv(path, chunksize=chunk_size, **kwargs):
        chunks.append(chunk)
        rows_read += len(chunk)
        on_progress(min(rows_read, total), total)
    if not chunks:
        on_progress(0, total)
        return pd.DataFrame()
    df = pd.concat(chunks, ignore_index=True)
    on_progress(min(len(df), total), total)
    return df


def _read_excel_with_progress(path, sheet, hr, total, on_progress):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        target = sheet if (sheet and sheet in names) else names[0]
        ws = wb[target]
        header_vals = next(ws.iter_rows(min_row=hr + 1, max_row=hr + 1, values_only=True), None)
        if header_vals is None:
            on_progress(0, total or 1)
            return pd.DataFrame()
        columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(header_vals)]
        rows, rows_read = [], 0
        emit_every = max(500, (total or 1) // 50)
        for row in ws.iter_rows(min_row=hr + 2, values_only=True):
            rows.append(row)
            rows_read += 1
            if rows_read % emit_every == 0:
                on_progress(min(rows_read, total), total)
        # `pd.read_excel` drops trailing all-empty rows; `openpyxl.iter_rows`
        # keeps them. Mirror pandas so downstream blocks see the same shape.
        while rows and all(v is None or (isinstance(v, str) and v == "") for v in rows[-1]):
            rows.pop()
        on_progress(min(len(rows), total or len(rows) or 1), total or len(rows) or 1)
        return pd.DataFrame(rows, columns=columns)
    finally:
        wb.close()


def _read_source(
    path,
    sheet,
    header_row: int,
    data: dict | None = None,
    *,
    on_progress=None,
) -> pd.DataFrame:
    ext = path.suffix.lower()
    hr = int(header_row or 0)
    total = _count_source_data_rows(path, sheet, hr, data) if on_progress else 0

    if on_progress:
        on_progress(0, total or 1)

    if ext in _EXCEL_EXT:
        if on_progress and total > 0:
            return _read_excel_with_progress(path, sheet, hr, total, on_progress)
        return pd.read_excel(path, sheet_name=(sheet or 0), skiprows=hr)
    if ext == ".parquet":
        df = pd.read_parquet(path)
        if on_progress:
            n = len(df)
            on_progress(n, n or 1)
        return df
    opts, _ = _resolve_read_options(path, data or {})
    if on_progress and total > 0:
        return _read_csv_with_progress(path, hr, opts, ext, total, on_progress)
    if ext in (".tsv", ".txt"):
        return pd.read_csv(path, sep="\t", skiprows=hr, **opts)
    return pd.read_csv(path, sep=None, engine="python", skiprows=hr, **opts)


def _dtype_label(dt) -> str:
    s = str(dt)
    if s.startswith("int") or s.startswith("Int") or s.startswith("uint"):
        return "BIGINT"
    if s.startswith("float") or s.startswith("Float"):
        return "DOUBLE"
    if s.startswith("bool"):
        return "BOOLEAN"
    if "datetime" in s:
        return "TIMESTAMP"
    return "VARCHAR"


def peek_source(
    pid: str, file: str, sheet: str | None, header_row, data: dict | None = None
) -> dict:
    path = storage.files_dir(pid) / file
    if not path.exists():
        raise ValueError(f"Fichier introuvable : {file}")
    sheets = None
    hr = int(header_row or 0)
    detected: dict = {}
    warnings: list[str] = []  # A.8 — pièges détectés (cellules fusionnées, etc.)
    if path.suffix.lower() in _EXCEL_EXT:
        with pd.ExcelFile(path) as xl:  # context-manager closes the file handle
            sheets = list(xl.sheet_names)
            target = sheet if (sheet and sheet in sheets) else sheets[0]
            df = pd.read_excel(xl, sheet_name=target, skiprows=hr, nrows=100)
        warnings.extend(_xlsx_warnings(path, target))
    else:
        opts, detected = _resolve_read_options(path, data or {})
        df = _read_source(path, sheet, hr, data).head(100)
        # Surface the effective sep too, so the aperçu can name it (A.3).
        # pandas with sep=None+engine='python' sniffs internally — peek again
        # with the same options to recover what it picked.
        try:
            with path.open("r", encoding=opts.get("encoding") or "utf-8", errors="replace") as f:
                head = "".join([next(f, "") for _ in range(8)])
            detected["sep"] = _guess_sep_from_head(head)
        except Exception:  # noqa: BLE001
            pass
    columns = [{"name": str(c), "type": _dtype_label(df[c].dtype)} for c in df.columns]
    return {"sheets": sheets, "columns": columns, "detected": detected, "warnings": warnings}


def _xlsx_warnings(path, sheet: str | None) -> list[str]:
    """Détecte les pièges d'un xlsx FR — cellules fusionnées dans la zone
    d'en-tête (cf. todo A.8). Best-effort : un fichier qu'openpyxl ne sait pas
    ouvrir retourne juste [], on n'interrompt pas l'aperçu pour ça."""
    out: list[str] = []
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        try:
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            # Fusion dans les 10 premières lignes = en-tête, où ça produit des
            # cellules vides à l'aperçu sans qu'on sache pourquoi.
            merged_in_header = [r for r in ws.merged_cells.ranges if r.min_row <= 10]
            if merged_in_header:
                out.append(
                    f"{len(merged_in_header)} cellule(s) fusionnée(s) dans les 10 premières "
                    "lignes. pandas vide les fusions — l'aperçu peut afficher des trous qui "
                    "n'existent pas dans Excel. Ouvrez le fichier et défusionnez avant import."
                )
            # En-tête potentiellement sur plusieurs lignes : la 1ère ligne a peu
            # de valeurs (uniquement les cellules de gauche), la 2e en a plus.
            row1 = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1), [])]
            row2 = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2), [])]
            n1 = sum(1 for v in row1 if v not in (None, ""))
            n2 = sum(1 for v in row2 if v not in (None, ""))
            if n1 > 0 and n2 > n1 + 1:
                out.append(
                    "L'en-tête semble s'étendre sur plusieurs lignes (1re ligne incomplète). "
                    "Augmentez « Lignes d'en-tête à ignorer » pour démarrer à la bonne ligne."
                )
        finally:
            wb.close()
    except Exception:  # noqa: BLE001 — best-effort
        pass
    return out


def _guess_sep_from_head(text: str) -> str | None:
    """Best-effort separator guess from a few header/data lines (CSV-only)."""
    lines = [ln for ln in text.splitlines() if ln.strip()][:8]
    if not lines:
        return None
    best, best_score = None, 0
    for sep in (";", ",", "\t", "|"):
        counts = [ln.count(sep) for ln in lines]
        if counts[0] == 0:
            continue
        # Stable count across lines = strong signal.
        if all(c == counts[0] for c in counts):
            score = counts[0] * 10
        else:
            score = counts[0]
        if score > best_score:
            best, best_score = sep, score
    return best


# --------------------------------------------------------------------------- #
# Graph ordering
# --------------------------------------------------------------------------- #
def _topo_order(nodes: dict, edges: list) -> list[str]:
    indeg = {nid: 0 for nid in nodes}
    succ = {nid: [] for nid in nodes}
    for e in edges:
        s, t = e.get("source"), e.get("target")
        if s in nodes and t in nodes:
            succ[s].append(t)
            indeg[t] += 1
    queue = [nid for nid in nodes if indeg[nid] == 0]
    order = []
    while queue:
        n = queue.pop(0)
        order.append(n)
        for m in succ[n]:
            indeg[m] -= 1
            if indeg[m] == 0:
                queue.append(m)
    # append any nodes left out by a cycle so we never silently drop work
    for nid in nodes:
        if nid not in order:
            order.append(nid)
    return order


def _ancestors(node_id: str, edges: list) -> set[str]:
    preds = {}
    for e in edges:
        preds.setdefault(e.get("target"), []).append(e.get("source"))
    seen = set()
    stack = [node_id]
    while stack:
        n = stack.pop()
        for p in preds.get(n, []):
            if p and p not in seen:
                seen.add(p)
                stack.append(p)
    return seen


# --------------------------------------------------------------------------- #
# Inputs
# --------------------------------------------------------------------------- #
def _node_inputs(pid, wid, node, edges) -> list[tuple[str, pd.DataFrame]]:
    ins = []
    for e in edges:
        if e.get("target") != node["id"]:
            continue
        src = e.get("source")
        sh = e.get("sourceHandle") or "out"
        alias = e.get("targetHandle") or "in"
        path = storage.node_parquet(pid, wid, src, sh)
        if not path.exists():
            # Either upstream wasn't executed yet, or the edge points to a
            # handle that no longer exists on the source block (an output was
            # removed but the edge stayed). The second case is silent fuel for
            # ghost data — name the handle so it's findable.
            raise ValueError(
                f"une entrée n'a pas encore été exécutée (source : {src} / sortie : {sh})"
            )
        ins.append((alias, pd.read_parquet(path)))
    return ins


def _single(ins):
    if not ins:
        raise ValueError("aucune entrée connectée")
    return ins[0][1]


# --------------------------------------------------------------------------- #
# Block runners  ->  (outputs: {handle: DataFrame}, extra_meta: dict)
# --------------------------------------------------------------------------- #
def _run_source(con, pid, node, ins, *, on_progress=None):
    d = node["data"]
    file = d.get("file")
    if not file:
        raise ValueError("aucun fichier sélectionné")
    path = storage.files_dir(pid) / file
    if not path.exists():
        raise ValueError(f"fichier introuvable : {file}")
    df = _read_source(path, d.get("sheet"), d.get("header_row") or 0, d, on_progress=on_progress)
    # A.5 — casts utilisateur explicites (number / date / text / boolean) AVANT
    # matérialisation, pour ne pas forcer un détour par un bloc Nettoyage.
    df, casts_report = _apply_casts(df, d.get("casts"), d.get("decimal") or "", d.get("thousands"))
    # Avant matérialisation, on désamorce les colonnes object hétérogènes que
    # pyarrow refuserait de sérialiser. On le fait APRÈS les casts pour ne pas
    # masquer une colonne que l'utilisateur a déjà forcée en nombre/date.
    mixed_report = _normalize_mixed_object_columns(df)
    extra = {}
    if casts_report:
        extra["casts_report"] = casts_report
    if mixed_report:
        extra["mixed_types"] = mixed_report
    fp = _source_fingerprint(pid, node)
    if fp:
        extra["source_fingerprint"] = fp
    return {"out": df}, extra


def _open_sandboxed_duckdb():
    """B.4 — connexion DuckDB jetable pour exécuter du SQL utilisateur (mode raw).

    `enable_external_access = false` coupe en un seul interrupteur : accès
    filesystem, réseau, chargement d'extensions, ATTACH de bases externes. Le
    réglage est *one-way* (impossible à remettre à true dans la même session)
    — donc même un SQL malicieux qui essaie `SET enable_external_access = true`
    en début de requête est rejeté par DuckDB. Les fonctions interdites
    (`read_csv_auto`, `read_parquet`, `COPY ... TO`, `httpfs`, `INSTALL`,
    `LOAD`, `ATTACH`, `EXPORT DATABASE`) lèvent une `IOException`.

    On utilise quand même `register()` sur des DataFrames pandas en mémoire :
    ça ne passe pas par le filesystem, c'est autorisé après le verrouillage.
    """
    sandbox = duckdb.connect()
    sandbox.execute("SET enable_external_access = false")
    return sandbox


def _run_sql(con, pid, node, ins):
    d = node["data"]
    if not ins:
        raise ValueError("aucune entrée connectée")
    is_raw = d.get("mode") == "raw"
    # Mode raw → connexion sandboxée jetable (B.4). Mode builder → la connexion
    # partagée du workflow (le SQL est généré par nous, donc sûr).
    sql_con = _open_sandboxed_duckdb() if is_raw else con
    aliases = []
    try:
        for alias, df in ins:
            sql_con.register(alias, df)
            aliases.append(alias)
        primary = "in1" if "in1" in aliases else aliases[0]
        if is_raw:
            sql = (d.get("raw_sql") or "").strip()
            if not sql:
                raise ValueError("SQL brut vide")
        else:
            sql = query_builder.compile_query(d.get("query") or {}, primary)
        try:
            df = sql_con.execute(sql).df()
        except Exception as e:
            # Message dédié pour le cas sandbox : on dit clairement que c'est
            # une restriction de sécurité, pas un bug DuckDB. DuckDB ne renvoie
            # pas un message uniforme — on en reconnaît plusieurs formes.
            if is_raw:
                emsg = str(e).lower()
                sandbox_signals = (
                    "external access",
                    "external_access",
                    "file system operations are disabled",
                    "disabled by configuration",
                )
                if any(s in emsg for s in sandbox_signals):
                    raise ValueError(
                        "SQL brut refusé : accès filesystem/réseau désactivé dans cette "
                        "sandbox (read_csv_auto, COPY, httpfs, ATTACH… sont interdits). "
                        "Branchez vos fichiers via un bloc Source en amont. "
                        f"— Détail DuckDB : {e}"
                    )
            raise ValueError(f"{e}  —  SQL généré : {sql}")
    finally:
        # Sandbox jetable : `unregister` n'est pas requis (close suffit), mais
        # sur la connexion partagée on doit nettoyer pour ne pas polluer les
        # blocs aval qui réutilisent les mêmes alias.
        if is_raw:
            sql_con.close()
        else:
            for alias in aliases:
                con.unregister(alias)
    return {"out": df}, {"compiled_sql": sql}


def _run_dedup(con, pid, node, ins):
    d = node["data"]
    df = _single(ins)
    # Explicitly-selected key columns must exist in the actual input. Silently
    # dropping a missing one would make dedup fall back to fewer keys (e.g. the
    # first column only) and produce wrong-but-plausible groups with no warning.
    requested = list(d.get("key_columns") or [])
    missing = [k for k in requested if k not in df.columns]
    if missing:
        raise ValueError(
            "colonnes-clés introuvables dans l'entrée : "
            + ", ".join(missing)
            + " — réexécutez le bloc amont, puis re-cochez les colonnes-clés."
        )
    keys = requested or list(df.columns)
    keep = d.get("keep", "first")
    if keep not in ("first", "last"):
        keep = "first"
    dups_mode = d.get("dups_mode", "all")

    dup_any = df.duplicated(subset=keys, keep=False)
    exemplar = ~df.duplicated(subset=keys, keep=keep)  # one row per group
    kept = df[exemplar]
    uniques = df[~dup_any]
    if dups_mode == "exemplar":
        dups = df[dup_any & exemplar]
    elif dups_mode == "extra":
        dups = df[df.duplicated(subset=keys, keep=keep)]
    else:
        dups = df[dup_any]
    # Group identical-key rows together so the duplicate groups are obvious when
    # the user inspects the "Doublons" output (scattered rows look unrelated).
    try:
        dups = dups.sort_values(by=keys, kind="stable")
    except TypeError:
        pass  # unorderable mixed types
    return {
        "kept": kept.reset_index(drop=True),
        "dups": dups.reset_index(drop=True),
        "uniques": uniques.reset_index(drop=True),
    }, {}


def _run_validate(con, pid, node, ins):
    d = node["data"]
    df = _single(ins)
    if d.get("mode") == "route":
        return _run_route(d, df)
    col = d.get("target_column")
    if not col or col not in df.columns:
        raise ValueError("choisissez la colonne à contrôler")
    flags = []
    for v in df[col].tolist():
        ok, _ = _validate_one(v, d)
        flags.append(ok)
    out = df.copy()
    if d.get("add_flag"):
        out["valide"] = ["oui" if f else "non" for f in flags]
    mask = pd.Series(flags, index=out.index)
    valid_df = out[mask].reset_index(drop=True)
    invalid_df = out[~mask].reset_index(drop=True)
    n_in = len(df)
    n_out = len(valid_df) + len(invalid_df)
    return {
        "valid": valid_df,
        "invalid": invalid_df,
    }, {"input_rows": n_in, "unrouted_rows": max(0, n_in - n_out)}


# --------------------------------------------------------------------------- #
# Routing (neutral multi-output classification, set-algebra conditions)
# --------------------------------------------------------------------------- #
_VECTORIZABLE_TESTS = frozenset(
    {
        "equals",
        "not_equals",
        "is_empty",
        "not_empty",
        "starts_with",
        "ends_with",
        "contains",
        "not_contains",
        "length_eq",
        "length_min",
        "length_max",
        "is_in",
        "regex",
        "regex_full",
    }
)
# String tests whose right-hand side can be either a literal or another column.
_VS_COLUMN_STR_TESTS = frozenset(
    {
        "equals",
        "not_equals",
        "contains",
        "not_contains",
        "starts_with",
        "ends_with",
    }
)
_NUMERIC_TESTS = frozenset({"num_eq", "num_ne", "num_gt", "num_ge", "num_lt", "num_le"})


# Tests whose negation has a positive counterpart : utile en mode multi-valeur
# pour interpréter « ne contient aucune de ces valeurs » comme NOT(contient une).
_NEGATION_OF = {"not_contains": "contains", "not_equals": "equals"}


def _multi_values(rule: dict) -> list | None:
    """Renvoie la liste des valeurs `rule.values` si elle est non-vide (après
    strip), sinon None. Permet de dispatcher en mode multi-valeur : un même test
    contre N valeurs, combiné en OR (ANY-match)."""
    vals = rule.get("values")
    if not isinstance(vals, list):
        return None
    cleaned = [v for v in vals if v is not None and str(v).strip() != ""]
    return cleaned if cleaned else None


def _rule_mask(df: pd.DataFrame, rule: dict, cs: bool, default_col: str) -> pd.Series:
    """Boolean Series: does each row satisfy this single rule (optionally negated)?
    Vectorized via pandas .str.* / numeric ops for the common tests; per-row
    _rule_ok fallback for positional/numeric-string tests. A rule may compare
    against either a literal (rule['value']) or another column (rule['via']
    == 'column' + rule['column2']). En mode multi-valeur (`rule.values` =
    liste), la règle matche si au moins une valeur de la liste matche le test
    (pour les tests positifs) ; pour `not_contains`/`not_equals`, on bascule
    sur l'équivalent positif et on inverse (« ne matche aucune »)."""
    # Multi-valeur : recurse sur chaque valeur avec le test (positif si négation)
    # et combine en OR. La `negate` finale est appliquée à la fin pour ne pas
    # se télescoper avec la conversion not_X → X.
    mv = _multi_values(rule)
    if mv is not None:
        is_neg = rule.get("test") in _NEGATION_OF
        pos_test = _NEGATION_OF.get(rule.get("test"), rule.get("test"))
        sub_base = {k: v for k, v in rule.items() if k not in ("values", "negate")}
        sub_base["test"] = pos_test
        masks = [_rule_mask(df, {**sub_base, "value": v}, cs, default_col) for v in mv]
        m = masks[0]
        for x in masks[1:]:
            m = m | x
        if is_neg:
            m = ~m
        m = m.astype(bool)
        return ~m if rule.get("negate") else m

    col = rule.get("column") or default_col
    if not col or col not in df.columns:
        return pd.Series(False, index=df.index)
    test = rule.get("test")
    # vs-column mode is sticky: when explicitly chosen, an unset/invalid column2
    # makes the rule match nothing (rather than silently degrading to the
    # literal r.value, which would surprise the user).
    if rule.get("via") == "column":
        if rule.get("column2") not in df.columns:
            return pd.Series(False, index=df.index)
        via_col = True
    else:
        via_col = False

    # Numeric comparisons (==, !=, >, >=, <, <=) — coerce both sides; rows where
    # either side fails to parse are non-matching (and never raise).
    if test in _NUMERIC_TESTS:
        a = pd.to_numeric(df[col], errors="coerce")
        if via_col:
            b = pd.to_numeric(df[rule["column2"]], errors="coerce")
        else:
            try:
                b = float(rule.get("value"))
            except (TypeError, ValueError):
                m = pd.Series(False, index=df.index)
                m = m.astype(bool)
                return ~m if rule.get("negate") else m
        op = {
            "num_eq": a.eq,
            "num_ne": a.ne,
            "num_gt": a.gt,
            "num_ge": a.ge,
            "num_lt": a.lt,
            "num_le": a.le,
        }[test]
        m = op(b).fillna(False).astype(bool)
        return ~m if rule.get("negate") else m

    if test in _VECTORIZABLE_TESTS:
        s = df[col].astype("string").fillna("")
        ss = s if cs else s.str.lower()
        # Right-hand side: column (per-row) or literal.
        if via_col and test in _VS_COLUMN_STR_TESTS:
            s2 = df[rule["column2"]].astype("string").fillna("")
            ss2 = s2 if cs else s2.str.lower()
            if test == "equals":
                m = ss == ss2
            elif test == "not_equals":
                m = ss != ss2
            elif test == "starts_with":
                # No vectorized pandas op for "row.A starts with row.B"; the
                # zip+gen-expr is still ~20× faster than a .apply with isinstance.
                m = pd.Series([a.startswith(b) for a, b in zip(ss, ss2)], index=df.index)
            elif test == "ends_with":
                m = pd.Series([a.endswith(b) for a, b in zip(ss, ss2)], index=df.index)
            elif test == "contains":
                m = pd.Series([b in a for a, b in zip(ss, ss2)], index=df.index)
            else:  # not_contains
                m = pd.Series([b not in a for a, b in zip(ss, ss2)], index=df.index)
            m = m.astype(bool)
            return ~m if rule.get("negate") else m
        raw = rule.get("value")
        vv = "" if raw is None else (str(raw) if cs else str(raw).lower())
        if test == "equals":
            m = ss == vv
        elif test == "not_equals":
            m = ss != vv
        elif test == "is_empty":
            m = s == ""
        elif test == "not_empty":
            m = s != ""
        elif test == "starts_with":
            m = ss.str.startswith(vv, na=False)
        elif test == "ends_with":
            m = ss.str.endswith(vv, na=False)
        elif test == "contains":
            m = ss.str.contains(vv, regex=False, na=False)
        elif test == "not_contains":
            m = ~ss.str.contains(vv, regex=False, na=False)
        elif test == "is_in":
            items = [x.strip() for x in str(raw or "").split(",") if x.strip() != ""]
            if not cs:
                items = [x.lower() for x in items]
            m = ss.isin(items)
        elif test in ("length_eq", "length_min", "length_max"):
            try:
                n = int(raw)
            except (TypeError, ValueError):
                m = pd.Series(False, index=df.index)
            else:
                lengths = s.str.len()
                m = (
                    lengths == n
                    if test == "length_eq"
                    else lengths >= n
                    if test == "length_min"
                    else lengths <= n
                )
        else:  # regex / regex_full
            try:
                if test == "regex":
                    m = s.str.contains(str(raw or ""), regex=True, na=False, case=cs)
                else:
                    m = s.str.fullmatch(str(raw or ""), case=cs).fillna(False)
            except re.error:
                m = pd.Series(False, index=df.index)
    else:
        m = df[col].map(lambda v: _rule_ok("" if _is_null(v) else str(v), rule, cs))
    m = m.astype(bool)
    return ~m if rule.get("negate") else m


def _eval_condition(df: pd.DataFrame, condition: dict, cs: bool, default_col: str) -> pd.Series:
    """Evaluate a DNF condition over a DataFrame -> boolean Series.
    condition = {groups: [{rules: [...]}, ...]} : OR of groups, each group is the
    AND of its rules. Set algebra: union of intersections (complement via negate).
    No groups (or only empty groups) -> matches nothing."""
    groups = (condition or {}).get("groups") or []
    result = pd.Series(False, index=df.index)
    for g in groups:
        rules = g.get("rules") or []
        if not rules:
            continue
        gm = pd.Series(True, index=df.index)
        for r in rules:
            gm &= _rule_mask(df, r, cs, default_col)
        result |= gm
    return result


# --------------------------------------------------------------------------- #
# Calcul block — "value_when" group function (pandas-side)
# --------------------------------------------------------------------------- #
_VALUE_WHEN_REDUCERS = frozenset({"first", "last", "min", "max", "sum", "avg", "concat"})


def _vw_reduce(s, reducer):
    """Reduce a Series to a single value, ignoring NaN.
    Returns NA when no value matched (so the fallback kicks in) — sum is special-
    cased because pandas would otherwise turn an all-NaN series into 0."""
    if not s.notna().any():
        return pd.NA
    if reducer == "first":
        return s.dropna().iloc[0]
    if reducer == "last":
        return s.dropna().iloc[-1]
    if reducer == "concat":
        return ", ".join(str(x) for x in s.dropna())
    n = pd.to_numeric(s, errors="coerce")
    if not n.notna().any():
        return pd.NA
    if reducer == "min":
        return n.min()
    if reducer == "max":
        return n.max()
    if reducer == "sum":
        return n.sum()
    if reducer == "avg":
        return n.mean()
    return pd.NA


def _apply_value_when(con, df: pd.DataFrame, group: dict) -> pd.DataFrame:
    """Compute pandas-only 'value_when' group columns on top of df.
    Each function picks a per-row value (a column or a compiled formula), keeps
    only the rows whose condition matches, reduces per group (first/last follow
    group.order_by ; min/max/sum/avg/concat are order-independent), then
    broadcasts the result to every row of the group. A group with no matching
    row falls back to NULL or to a fixed user value."""
    g = group or {}
    funcs = [
        f
        for f in (g.get("functions") or [])
        if f.get("fn") == "value_when" and f.get("enabled") is not False
    ]
    if not funcs:
        return df
    keys = [c for c in (g.get("partition_by") or []) if c in df.columns]
    order_pairs = []
    for o in g.get("order_by") or []:
        c = o.get("column") if isinstance(o, dict) else o
        if c and c in df.columns:
            order_pairs.append((c, bool(isinstance(o, dict) and o.get("desc"))))
    if order_pairs:
        sort_cols = [c for c, _ in order_pairs]
        ascending = [not d for _, d in order_pairs]
        sorted_idx = df.sort_values(sort_cols, ascending=ascending, kind="stable").index
    else:
        sorted_idx = df.index

    out = df.copy()
    for f in funcs:
        name = (f.get("name") or "").strip()
        if not name:
            continue
        src = f.get("source") or {}
        if src.get("kind") == "formula":
            expr = (src.get("expr") or "").strip()
            if not expr:
                continue
            try:
                sql_expr = compile_formula(expr)
            except FormulaError:
                continue
            con.register("_vw_in", out)
            try:
                values = con.execute(f"SELECT ({sql_expr}) AS _v FROM _vw_in").df()["_v"]
            finally:
                con.unregister("_vw_in")
            values.index = out.index
        else:
            col = src.get("column")
            if not col or col not in out.columns:
                continue
            values = out[col]

        mask = _eval_condition(out, f.get("condition") or {}, True, None).astype(bool)
        picked = values.where(mask)  # non-matching rows -> NaN
        reducer = f.get("reducer") or "first"
        if reducer not in _VALUE_WHEN_REDUCERS:
            reducer = "first"
        # first/last need the user's sort order ; the others don't.
        picked_s = picked.loc[sorted_idx] if reducer in ("first", "last") else picked
        if keys:
            key_series = [out[k].loc[picked_s.index] for k in keys]
            grouped = picked_s.groupby(key_series, dropna=False)
        else:
            grouped = picked_s.groupby(lambda _: 0, dropna=False)
        agg = grouped.transform(lambda s: _vw_reduce(s, reducer)).reindex(out.index)

        fb = f.get("fallback") or {}
        if fb.get("mode") == "value":
            agg = agg.where(agg.notna(), fb.get("value"))

        out[name] = agg
    return out


def _group_check_core(work, df, keys, ck, cs, default_col):
    """Per-row mask for ONE *property* group check (no WHEN premise), or None when
    the check is incomplete (no column / no N / no value yet) so it can be skipped
    during editing. The premise + the 'rows_satisfy' consequent live one level up,
    in _group_one_check_mask, which calls this on the matching subset of rows.

    Group-size checks (no column needed):
        size_eq/size_min/size_max : the group has = / ≥ / ≤ N rows.
    Column checks (need a column):
        unique       : the value differs from every other in its group.
        constant     : the group holds a single distinct value.
        distinct_eq/min/max : the group has = / ≥ / ≤ N distinct values.
        contains/not_contains : the group does / doesn't hold a given value.
        no_null      : the column is never null/empty anywhere in the group.
        not_all_null : the column has at least one non-null value in the group.
        all_null     : the column is empty in every row of the group (inverse of no_null)."""
    check = ck.get("check") or "unique"

    def _int(x):
        try:
            return int(x)
        except (TypeError, ValueError):
            return None

    if check in ("size_eq", "size_min", "size_max"):
        k = _int(ck.get("n"))
        if k is None:
            return None
        sizes = (
            df.groupby(keys, dropna=False)[keys[0]].transform("size")
            if keys
            else pd.Series(len(df), index=df.index)
        )
        if check == "size_eq":
            return sizes == k
        if check == "size_min":
            return sizes >= k
        return sizes <= k

    col = ck.get("column") or default_col
    if not col:
        return None
    if col not in df.columns:
        raise ValueError(f"Contrôle par groupe : colonne « {col} » introuvable dans l'entrée.")

    def _distinct():
        if keys:
            return work.groupby(keys, dropna=False)[col].transform(
                lambda s: s.nunique(dropna=False)
            )
        return pd.Series(work[col].nunique(dropna=False), index=df.index)

    if check == "no_null":  # group OK iff col never null inside it
        if keys:
            return (
                work.groupby(keys, dropna=False)[col]
                .transform(lambda s: s.notna().all())
                .astype(bool)
            )
        return pd.Series(bool(work[col].notna().all()), index=df.index)
    if check == "not_all_null":  # group OK iff col not entirely null
        if keys:
            return (
                work.groupby(keys, dropna=False)[col]
                .transform(lambda s: s.notna().any())
                .astype(bool)
            )
        return pd.Series(bool(work[col].notna().any()), index=df.index)
    if check == "all_null":  # group OK iff col empty everywhere (inverse of no_null)
        if keys:
            return (
                work.groupby(keys, dropna=False)[col]
                .transform(lambda s: s.isna().all())
                .astype(bool)
            )
        return pd.Series(bool(work[col].isna().all()), index=df.index)
    if check == "unique":
        # group-atomic: the WHOLE group is OK iff every value is distinct
        # (nunique == size), broadcast to all its rows — so a group is never split.
        if keys:
            g = work.groupby(keys, dropna=False)[col]
            return g.transform(lambda s: s.nunique(dropna=False)) == g.transform("size")
        return pd.Series(work[col].nunique(dropna=False) == len(work), index=df.index)
    if check == "constant":
        return _distinct() <= 1
    if check in ("distinct_eq", "distinct_min", "distinct_max"):
        k = _int(ck.get("n"))
        if k is None:
            return None
        nun = _distinct()
        if check == "distinct_eq":
            return nun == k
        if check == "distinct_min":
            return nun >= k
        return nun <= k
    if check in ("contains", "not_contains"):
        # vs-column mode: per-row "col contains col2" check, then "any" per group.
        if ck.get("via") == "column":
            col2 = ck.get("column2")
            if not col2 or col2 not in df.columns:
                return None
            a = work[col].astype("string").fillna("")
            b = work[col2].astype("string").fillna("")
            if not cs:
                a, b = a.str.lower(), b.str.lower()
            per_row = pd.Series([bb in aa for aa, bb in zip(a, b)], index=df.index)
            if keys:
                has = per_row.groupby([df[k] for k in keys], dropna=False).transform("any")
            else:
                has = pd.Series(bool(per_row.any()), index=df.index)
            has = has.astype(bool)
            return has if check == "contains" else ~has
        target = ck.get("value")
        if target is None or target == "":
            return None
        if not cs and isinstance(target, str):
            target = target.lower()
        if keys:
            has = work.groupby(keys, dropna=False)[col].transform(lambda s: (s == target).any())
        else:
            has = pd.Series(bool((work[col] == target).any()), index=df.index)
        has = has.astype(bool)
        return has if check == "contains" else ~has

    # --- two-column checks: compare the checked column to a second one --------
    if check == "rows_cols_cmp":
        # Per-row "colA <op> colB" verdict, then "all" over the group — group OK
        # iff every row passes. eq/ne are string compares (casefold for ci);
        # gt/ge/lt/le coerce to numeric (non-coercible rows fail).
        col2 = ck.get("column2")
        if not col2 or col2 not in df.columns:
            return None
        op = ck.get("op") or "eq"
        if op in ("eq", "ne"):
            a = work[col].astype("string").fillna("")
            b = work[col2].astype("string").fillna("")
            per_row = (a == b) if op == "eq" else (a != b)
        else:
            a = pd.to_numeric(df[col], errors="coerce")
            b = pd.to_numeric(df[col2], errors="coerce")
            per_row = {"lt": a < b, "le": a <= b, "gt": a > b, "ge": a >= b}[op]
            per_row = per_row.fillna(False)
        if keys:
            return (
                per_row.groupby([df[k] for k in keys], dropna=False).transform("all").astype(bool)
            )
        return pd.Series(bool(per_row.all()), index=df.index)
    if check in ("distinct_cmp", "determines"):
        col2 = ck.get("column2")
        if not col2:
            return None
        if col2 not in df.columns:
            raise ValueError(f"Contrôle par groupe : colonne « {col2} » introuvable dans l'entrée.")
        nun = (
            (
                lambda c: work.groupby(keys, dropna=False)[c].transform(
                    lambda s: s.nunique(dropna=False)
                )
            )
            if keys
            else (lambda c: pd.Series(work[c].nunique(dropna=False), index=df.index))
        )
        if check == "distinct_cmp":  # nb distinct(col) <op> nb distinct(col2)
            na, nb = nun(col), nun(col2)
            op = ck.get("op") or "eq"
            return {
                "eq": na == nb,
                "ne": na != nb,
                "lt": na < nb,
                "le": na <= nb,
                "gt": na > nb,
                "ge": na >= nb,
            }.get(op, na == nb)
        # determines: col → col2 is a function iff #distinct(col) == #distinct (col,col2)
        a = work[col].astype("string")
        pair = a.str.cat(work[col2].astype("string"), sep="\x1f", na_rep="\x00")
        tmp = work.assign(__a=a, __p=pair)
        if keys:
            g = tmp.groupby(keys, dropna=False)
            return g["__a"].transform(lambda s: s.nunique(dropna=False)) == g["__p"].transform(
                lambda s: s.nunique(dropna=False)
            )
        return pd.Series(
            tmp["__a"].nunique(dropna=False) == tmp["__p"].nunique(dropna=False), index=df.index
        )
    return None


def _nonempty_groups(dnf):
    """The groups of a rules-DNF that actually carry rules (ignore blanks left by
    the editor). Empty -> the condition is inert."""
    return [g for g in ((dnf or {}).get("groups") or []) if (g.get("rules") or [])]


def _eval_premise(df, when, cs, default_col):
    """Row mask for a check's optional WHEN premise, or None when there is no
    effective premise — so callers keep the fast, unfiltered path untouched."""
    groups = _nonempty_groups(when)
    if not groups:
        return None
    return _eval_condition(df, {"groups": groups}, cs, default_col).astype(bool)


def _broadcast_group_verdict(df, keys, sub_verdict):
    """Map an atomic per-group verdict (computed over a premise-matching SUBSET,
    constant within each group) back onto every row of df by key. Groups with no
    matching row are absent from the subset -> conforme by vacuity (True)."""
    sub = sub_verdict.astype(bool)
    if not keys:
        return pd.Series(bool(sub.all()) if len(sub) else True, index=df.index)
    gv = df.loc[sub.index, keys].assign(__ok=sub.to_numpy()).drop_duplicates(keys)
    ok = df[keys].merge(gv, on=keys, how="left")["__ok"]  # unique right keys -> left order kept;
    return pd.Series(
        ok.astype("boolean").fillna(True).to_numpy(dtype=bool), index=df.index
    )  # NaN (no match) -> True


def _group_one_check_mask(work, df, keys, ck, cs, default_col):
    """Per-row mask for ONE group check, atomic per group. Adds two things on top
    of _group_check_core:

      WHEN premise (ck['when'], a rules-DNF): the check applies only to the group's
        rows that satisfy it. The matching subset is verdicted by _group_check_core,
        then broadcast to the whole group (empty subset -> conforme by vacuity).
      rows_satisfy (ck['then'], a rules-DNF): every matching row must satisfy it; the
        group fails iff any matching row violates.

    Returns None when the check is incomplete (skipped during editing)."""
    check = ck.get("check") or "unique"
    pm = _eval_premise(df, ck.get("when"), cs, default_col)

    if check == "rows_satisfy":
        groups = _nonempty_groups(ck.get("then"))
        if not groups:
            return None
        cm = _eval_condition(df, {"groups": groups}, cs, default_col).astype(bool)
        viol = ~cm if pm is None else (pm & ~cm)
        if not keys:
            return pd.Series(not bool(viol.any()), index=df.index)
        bad = viol.groupby([df[k] for k in keys], dropna=False).transform("any").astype(bool)
        return ~bad

    if pm is None:  # no premise -> existing fast path
        return _group_check_core(work, df, keys, ck, cs, default_col)
    if not pm.any():  # premise never triggers -> all conforme
        return pd.Series(True, index=df.index)
    core = _group_check_core(work[pm], df[pm], keys, ck, cs, default_col)
    if core is None:
        return None
    return _broadcast_group_verdict(df, keys, core)


def _group_condition_mask(df: pd.DataFrame, cond: dict, cs: bool, default_col: str) -> pd.Series:
    """Per-row mask for a 'group' condition: group rows by the key columns, then
    apply one or several *checks* on the group / a checked column. A row satisfies
    the condition iff its group passes EVERY check (AND). Returns True for rows
    that satisfy (a 'NON' on the output isolates the offenders).

    Checks live in cond['checks'] (a list); legacy conditions with a single inline
    check (cond['check']/['column']/['n']/['value']) are supported transparently."""
    keys = [c for c in (cond.get("group_by") or []) if c in df.columns]
    checks = cond.get("checks")
    if not checks:  # legacy single-check shape
        checks = [
            {
                "check": cond.get("check"),
                "column": cond.get("column"),
                "n": cond.get("n"),
                "value": cond.get("value"),
            }
        ]

    # case-fold once over the keys and every checked column (for ci comparisons)
    cols_used = list(keys)
    for ck in checks:
        for c in (ck.get("column") or default_col, ck.get("column2")):
            if c and c in df.columns and c not in cols_used:
                cols_used.append(c)
    work = df
    if not cs:
        work = df.copy()
        for c in cols_used:
            if work[c].dtype == object:
                work[c] = work[c].map(lambda v: v.lower() if isinstance(v, str) else v)

    mask = pd.Series(True, index=df.index)
    applied = 0
    for ck in checks:
        m = _group_one_check_mask(work, df, keys, ck, cs, default_col)
        if m is not None:
            mask &= m.astype(bool)
            applied += 1
    if not applied:
        raise ValueError(
            f"Contrôle par groupe « {cond.get('name') or '?'} » : "
            "configurez au moins un contrôle (colonne / valeur / N)."
        )
    return mask


def _condition_mask(df: pd.DataFrame, cond: dict, cs: bool, default_col: str) -> pd.Series:
    """Boolean Series for a *named* condition object. Three flavours:
    - kind 'rules': a DNF (groups OR'd, rules AND'd) — delegates to _eval_condition.
    - kind 'mask' : a positional mask regex, full-matched on the column.
    - kind 'group': a whole-table group check (uniqueness/constancy within a key).
    A condition may override the column (cond['column']); otherwise default_col."""
    cond = cond or {}
    if cond.get("kind") == "group":
        return _group_condition_mask(df, cond, cs, default_col)
    col = cond.get("column") or default_col
    if cond.get("kind") == "mask":
        if not col or col not in df.columns:
            return pd.Series(False, index=df.index)
        pat = _mask_pattern(cond.get("segments") or [], cs)
        flags = 0 if cs else re.IGNORECASE

        def _m(v):
            s = "" if _is_null(v) else str(v)
            return bool(re.fullmatch(pat, s, flags)) if pat else (s == "")

        return df[col].map(_m).astype(bool)
    return _eval_condition(df, {"groups": cond.get("groups") or []}, cs, col)


def _output_mask(
    df: pd.DataFrame, o: dict, conds: dict, cs: bool, default_col: str, cache: dict | None = None
) -> pd.Series:
    """Row mask for one output. New shape: o['match'] = {conditionId, negate} points
    at a named condition. Legacy shape (kept working): an inline o['condition'] DNF.
    `cache` memoizes the (positive) per-condition mask across outputs — a group
    condition is reused by every output that references it (conforme + its NEG)."""
    match = o.get("match")
    if match and match.get("conditionId") in conds:
        cid = match["conditionId"]
        if cache is not None and cid in cache:
            m = cache[cid]
        else:
            m = _condition_mask(df, conds[cid], cs, default_col)
            if cache is not None:
                cache[cid] = m
        return ~m if match.get("negate") else m
    if o.get("condition") is not None:  # legacy inline DNF (pre-named-conditions)
        return _eval_condition(df, o.get("condition"), cs, default_col)
    return pd.Series(False, index=df.index)


def _augment_diagnostics(d: dict, df: pd.DataFrame, cs: bool, default_col: str) -> pd.DataFrame:
    """Conformity preset (intent 'control'): add the optional 'valide' column,
    computed against the first condition. No-op when it isn't requested."""
    if not d.get("add_flag"):
        return df
    conds = d.get("conditions") or []
    if not conds:
        return df
    cond = conds[0]
    if cond.get("kind") == "group":  # group checks route rows, they
        return df  # don't annotate a per-row column
    col = cond.get("column") or default_col
    if not col or col not in df.columns:
        return df
    out = df.copy()
    flags = []
    for v in out[col].tolist():
        s = "" if _is_null(v) else str(v)
        ok, _, _ = _condition_eval_str(s, cond, cs)
        flags.append(ok)
    out["valide"] = ["oui" if f else "non" for f in flags]
    return out


def _extract_key(series: pd.Series, extractor: dict, cs: bool) -> pd.Series:
    """Extract the 'part to watch' of each value as a string Series — the grouping
    key for the 'split by value' routing. Non-extractable values yield '' (→ else)."""
    s = series.astype("string").fillna("")
    ex = extractor or {}
    kind = ex.get("type", "after_last")
    if kind == "whole":
        # no extraction — the raw cell value is the grouping key
        key = s
    elif kind == "after_last":
        sep = ex.get("sep", ".")
        if sep:
            has = s.str.contains(re.escape(sep), regex=True)
            key = s.str.rpartition(sep)[2].where(has, "")  # part after last sep; none → ''
        else:
            key = s
    elif kind == "before_first":
        sep = ex.get("sep", "_")
        # part before the first separator; no separator → whole value
        key = s.str.partition(sep)[0] if sep else s
    elif kind == "segment":
        # Nth segment when split by a separator (e.g. a path level between '\').
        # index is 1-based; negative counts from the end (-1 = last). Out of range → ''.
        sep = ex.get("sep") or "\\"
        idx = int(ex.get("index") or 1)
        i = idx - 1 if idx > 0 else idx

        def _seg(v):
            parts = str(v).split(sep)
            return parts[i] if -len(parts) <= i < len(parts) else ""

        key = s.apply(_seg) if sep else s
    elif kind == "substring":
        start = max(1, int(ex.get("start") or 1)) - 1
        length = ex.get("length")
        end = start + int(length) if length not in (None, "", 0) else None
        key = s.str.slice(start, end)
    elif kind == "regex":
        pat = ex.get("pattern") or ""
        try:
            key = s.str.extract(f"({pat})" if "(" not in pat else pat, expand=True).iloc[:, 0]
        except re.error:
            key = pd.Series("", index=s.index, dtype="string")
        key = key.fillna("")
    else:
        key = s
    key = key.astype("string").fillna("")
    return key if cs else key.str.lower()


def _run_split(d: dict, df: pd.DataFrame):
    """Route rows by an extracted key: one output per distinct value (each output
    carries its 'value'); rows whose key matches none go to 'else'."""
    sp = d.get("split") or {}
    col = sp.get("column") or d.get("target_column")
    if not col or col not in df.columns:
        raise ValueError("choisissez la colonne à éclater")
    cs = bool(d.get("case_sensitive"))
    key = _extract_key(df[col], sp.get("extractor") or {}, cs)
    result, assigned = {}, pd.Series(False, index=df.index)
    for o in d.get("outputs") or []:
        oid = o.get("id")
        if not oid:
            continue
        val = o.get("value", "")
        val = val if cs else str(val).lower()
        m = (key == val) & ~assigned
        assigned |= m
        result[oid] = df[m].reset_index(drop=True)
    if d.get("else_enabled", True):
        result["else"] = df[~assigned].reset_index(drop=True)
    n_in = len(df)
    n_out = sum(len(v) for v in result.values())
    return result, {"input_rows": n_in, "unrouted_rows": max(0, n_in - n_out)}


def _run_route(d: dict, df: pd.DataFrame, diagnostics: bool = True):
    if (d.get("split") or {}).get("enabled"):
        return _run_split(d, df)
    cs = bool(d.get("case_sensitive"))
    default_col = d.get("target_column")
    routing = d.get("routing", "first")
    if diagnostics:
        df = _augment_diagnostics(d, df, cs, default_col)
    conds = {c.get("id"): c for c in (d.get("conditions") or []) if c.get("id")}
    result = {}
    assigned = pd.Series(False, index=df.index)
    cmask_cache: dict = {}  # share each named condition across outputs
    for o in d.get("outputs") or []:
        oid = o.get("id")
        if not oid:
            continue
        m = _output_mask(df, o, conds, cs, default_col, cache=cmask_cache)
        if routing == "first":  # partition: each row to the first matching output
            m = m & ~assigned
        assigned |= m
        result[oid] = df[m].reset_index(drop=True)
    if d.get("else_enabled", True):
        result["else"] = df[~assigned].reset_index(drop=True)
    n_in = len(df)
    n_out = sum(len(v) for v in result.values())
    return result, {"input_rows": n_in, "unrouted_rows": max(0, n_in - n_out)}


def route_preview(pid, wid, node_id, config: dict) -> dict:
    """Dry-run a routing config against the node's current input (no materialization).
    Returns per-output counts so the flow map can update live as the user edits."""
    wf = storage.get_workflow(pid, wid)
    if not wf:
        raise ValueError("Workflow introuvable")
    node = next((n for n in wf.get("nodes", []) if n["id"] == node_id), None)
    if not node:
        raise ValueError("Bloc introuvable")
    ins = _node_inputs(pid, wid, node, wf.get("edges", []))
    df = _single(ins)
    outs, _ = _run_route(config or {}, df, diagnostics=False)
    return {"total": int(len(df)), "counts": {h: int(len(v)) for h, v in outs.items()}}


def split_scan(pid, wid, node_id, config: dict) -> dict:
    """Discover the distinct extracted values of the 'split by value' config against
    the node's current input — feeds the 'Scan' that generates one output per value.
    Capped at 200 distinct values (truncated=True beyond)."""
    wf = storage.get_workflow(pid, wid)
    if not wf:
        raise ValueError("Workflow introuvable")
    node = next((n for n in wf.get("nodes", []) if n["id"] == node_id), None)
    if not node:
        raise ValueError("Bloc introuvable")
    ins = _node_inputs(pid, wid, node, wf.get("edges", []))
    df = _single(ins)
    sp = (config or {}).get("split") or {}
    col = sp.get("column") or (config or {}).get("target_column")
    if not col or col not in df.columns:
        raise ValueError("choisissez la colonne à éclater")
    cs = bool((config or {}).get("case_sensitive"))
    key = _extract_key(df[col], sp.get("extractor") or {}, cs)
    # Include the empty-key bucket: blank cells (or values that don't match the
    # extractor) become their own output, displayed as "(vide)" on the right.
    counts = key.value_counts()
    cap = 200
    values = [{"value": str(v), "count": int(c)} for v, c in counts.head(cap).items()]
    # a few raw → key examples so the user can immediately verify the extraction
    raw = df[col].astype("string").fillna("")
    seen, samples = set(), []
    for i in range(len(df)):
        r = str(raw.iloc[i])
        if r in seen:
            continue
        seen.add(r)
        samples.append({"raw": r[:160], "key": str(key.iloc[i])})
        if len(samples) >= 5:
            break
    return {
        "total": int(len(df)),
        "values": values,
        "samples": samples,
        "distinct": int(counts.size),
        "truncated": bool(counts.size > cap),
    }


def _run_pivot(con, pid, node, ins):
    d = node["data"]
    df = _single(ins)
    if d.get("mode") == "unpivot":
        vcs = [c for c in (d.get("value_columns") or []) if c in df.columns]
        if not vcs:
            raise ValueError("sélectionnez les colonnes à dépivoter")
        name_col = d.get("name_column") or "variable"
        val_col = d.get("value_column") or "valeur"
        id_vars = [c for c in df.columns if c not in vcs]
        out = df.melt(id_vars=id_vars, value_vars=vcs, var_name=name_col, value_name=val_col)
        return {"out": out}, {}
    # pivot
    idx = [c for c in (d.get("index_columns") or []) if c in df.columns]
    pc = d.get("pivot_column")
    vc = d.get("value_column")
    if not idx or not pc or not vc:
        raise ValueError("renseignez lignes, colonne à éclater et valeurs")
    aggmap = {
        "SUM": "sum",
        "AVG": "mean",
        "MIN": "min",
        "MAX": "max",
        "COUNT": "count",
        "MEDIAN": "median",
        "FIRST": "first",
        "LAST": "last",
    }
    fn = aggmap.get((d.get("agg") or "SUM").upper(), "sum")
    pt = pd.pivot_table(df, index=idx, columns=pc, values=vc, aggfunc=fn)
    pt = pt.reset_index()
    pt.columns = [str(c) for c in pt.columns]
    return {"out": pt}, {}


_CLEAN_LABELS = {
    "trim": "Supprimer les espaces",
    "upper": "MAJUSCULES",
    "lower": "minuscules",
    "capitalize": "Capitaliser",
    "replace": "Rechercher / remplacer",
    "fillna": "Remplir les vides",
    "to_number": "Convertir en nombre",
    "to_integer": "Convertir en entier",
    "to_date": "Convertir en date",
    "round": "Arrondir",
    "abs": "Valeur absolue",
}


def _run_clean(con, pid, node, ins):
    d = node["data"]
    df = _single(ins).copy()
    report = []
    for o in d.get("operations") or []:
        if o.get("enabled") is False:
            continue
        op = o.get("op")
        col = o.get("column")
        if not col or col not in df.columns:
            continue
        before = df[col]
        new, failed = _apply_clean(op, before, o)
        was_null = int(before.isna().sum())
        changed_mask = _changed_mask(before, new)
        changed = int(changed_mask.sum())
        samples = []
        for i in df.index[changed_mask][:5]:
            samples.append({"old": _json_safe(before.loc[i]), "new": _json_safe(new.loc[i])})
        df[col] = new
        report.append(
            {
                "op": op,
                "op_label": _CLEAN_LABELS.get(op, op),
                "column": col,
                "changed": changed,
                "failed": int(failed),
                "was_null": was_null,
                "samples": samples,
            }
        )
    return {"out": df}, {"clean_report": report}


def _changed_mask(old: pd.Series, new: pd.Series) -> pd.Series:
    o = old.astype(object).where(old.notna(), None)
    n = new.astype(object).where(new.notna(), None)
    return pd.Series([a != b for a, b in zip(o, n)], index=old.index)


def _apply_clean(op, s: pd.Series, o: dict):
    """Return (new_series, failed_count)."""
    failed = 0
    if op == "trim":
        return s.map(lambda x: x.strip() if isinstance(x, str) else x), 0
    if op == "upper":
        return s.map(lambda x: x.upper() if isinstance(x, str) else x), 0
    if op == "lower":
        return s.map(lambda x: x.lower() if isinstance(x, str) else x), 0
    if op == "capitalize":
        return s.map(lambda x: x.capitalize() if isinstance(x, str) else x), 0
    if op == "replace":
        find = o.get("find", "")
        repl = o.get("replace", "")
        use_re = bool(o.get("regex"))

        def rep(x):
            if not isinstance(x, str):
                return x
            return re.sub(find, repl, x) if use_re else x.replace(find, repl)

        return s.map(rep), 0
    if op == "fillna":
        val = o.get("value", "")
        if (o.get("value_type") or "text") == "number":
            try:
                val = float(val)
            except (TypeError, ValueError):
                val = 0
        return s.where(s.notna(), val), 0
    if op in ("to_number", "to_integer"):
        coerced = pd.to_numeric(
            s.map(lambda x: str(x).replace(",", ".") if isinstance(x, str) else x), errors="coerce"
        )
        failed = int((coerced.isna() & s.notna()).sum())
        if op == "to_integer":
            coerced = coerced.round().astype("Int64")
        return coerced, failed
    if op == "to_date":
        fmt = o.get("format") or None
        coerced = pd.to_datetime(s, format=fmt, errors="coerce")
        failed = int((coerced.isna() & s.notna()).sum())
        return coerced, failed
    if op == "round":
        digits = int(o.get("digits") or 0)
        coerced = pd.to_numeric(s, errors="coerce").round(digits)
        return coerced, int((coerced.isna() & s.notna()).sum())
    if op == "abs":
        coerced = pd.to_numeric(s, errors="coerce").abs()
        return coerced, int((coerced.isna() & s.notna()).sum())
    return s, 0


# Group/window functions: name -> whether it operates on a chosen target column.
_GROUP_NEEDS_COL = {
    "count_distinct",
    "is_duplicate",
    "sum",
    "avg",
    "min",
    "max",
    "median",
    "share",
    "first",
    "last",
    "prev",
    "next",
    "concat",
}


def _group_funcs_sql(group, df_cols) -> list[tuple[str, str]]:
    """Compile the Calcul block's "par groupe" functions into (name, sql_expr)
    pairs, one per added column. These are DuckDB window functions partitioned by
    the group key. Aggregates (count/sum/…) use a partition-only frame so they
    return the *whole group* value on every row (not a running total); ranking
    and offset functions (occurrence/first/last/prev/next) honour the order."""
    group = group or {}
    funcs = group.get("functions") or []
    if not funcs:
        return []
    part = [c for c in (group.get("partition_by") or []) if c in df_cols]
    order = []
    for o in group.get("order_by") or []:
        c = o.get("column") if isinstance(o, dict) else o
        if c and c in df_cols:
            order.append((c, bool(isinstance(o, dict) and o.get("desc"))))
    P = "PARTITION BY " + ", ".join(q_ident(c) for c in part) if part else ""
    O = (
        "ORDER BY " + ", ".join(q_ident(c) + (" DESC" if dsc else "") for c, dsc in order)
        if order
        else ""
    )

    def over(*clauses):
        return "OVER (" + " ".join(c for c in clauses if c) + ")"

    out: list[tuple[str, str]] = []
    for f in funcs:
        if f.get("enabled") is False:
            continue
        name = (f.get("name") or "").strip()
        fn = f.get("fn")
        col = f.get("column")
        if not name or not fn:
            continue
        if fn in _GROUP_NEEDS_COL and (not col or col not in df_cols):
            continue
        if fn == "value_when":  # pandas-side (see _apply_value_when)
            continue
        Y = q_ident(col) if col else None
        if fn == "occurrence":
            expr = f"ROW_NUMBER() {over(P, O)}"
        elif fn == "group_size":
            expr = f"COUNT(*) {over(P)}"
        elif fn == "count_distinct":
            expr = f"COUNT(DISTINCT {Y}) {over(P)}"
        elif fn == "is_duplicate":
            pxy = "PARTITION BY " + ", ".join(q_ident(c) for c in (part + [col]))
            expr = f"CASE WHEN COUNT(*) {over(pxy)} > 1 THEN 'oui' ELSE 'non' END"
        elif fn in ("sum", "avg", "min", "max", "median"):
            agg = {"sum": "SUM", "avg": "AVG", "min": "MIN", "max": "MAX", "median": "MEDIAN"}[fn]
            expr = f"{agg}({Y}) {over(P)}"
        elif fn == "share":
            expr = f"({Y} / NULLIF(SUM({Y}) {over(P)}, 0))"
        elif fn == "first":
            expr = f"first_value({Y}) {over(P, O)}"
        elif fn == "last":
            frame = "ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING"
            expr = f"last_value({Y}) {over(P, O, frame)}"
        elif fn == "prev":
            expr = f"lag({Y}) {over(P, O)}"
        elif fn == "next":
            expr = f"lead({Y}) {over(P, O)}"
        elif fn == "concat":
            expr = f"string_agg(CAST({Y} AS VARCHAR), ', ') {over(P)}"
        else:
            continue
        out.append((name, expr))
    return out


def _run_calc(con, pid, node, ins):
    d = node["data"]
    df = _single(ins)
    con.register("_calc_in", df)
    original = {str(c) for c in df.columns}
    known = set(original)
    try:
        # 1) group/window columns first, so formulas below can reference them.
        group_exprs = _group_funcs_sql(d.get("group"), original)
        if group_exprs:
            adds, replaces = [], []
            for name, expr in group_exprs:
                if name in original:
                    replaces.append(f"({expr}) AS {q_ident(name)}")
                else:
                    adds.append(f"({expr}) AS {q_ident(name)}")
                    known.add(name)
            sel = "SELECT *"
            if replaces:
                sel += " REPLACE (" + ", ".join(replaces) + ")"
            if adds:
                sel += ", " + ", ".join(adds)
            sql = sel + " FROM _calc_in"
        else:
            sql = "SELECT * FROM _calc_in"
        # 1b) "value_when" group columns — pandas-side because the rule engine
        # (Validation/Routage) is pandas, not SQL. We materialize what the SQL
        # step produced, apply the pandas transforms, then re-register so the
        # formula columns below see the new columns just like the SQL ones.
        group = d.get("group") or {}
        if any(
            f.get("fn") == "value_when" and f.get("enabled") is not False
            for f in (group.get("functions") or [])
        ):
            mat = _apply_value_when(con, con.execute(sql).df(), group)
            con.unregister("_calc_in")
            con.register("_calc_in", mat)
            known = {str(c) for c in mat.columns}
            sql = "SELECT * FROM _calc_in"
        # 2) Excel-like formula columns (layered on top of the group columns).
        for item in d.get("columns") or []:
            if item.get("enabled") is False:
                continue
            name = (item.get("name") or "").strip()
            expr = item.get("expr") or ""
            if not name or not expr.strip():
                continue
            sql_expr = compile_formula(expr)
            qname = q_ident(name)
            if name in known:
                sql = f"SELECT * REPLACE (({sql_expr}) AS {qname}) FROM ({sql}) q"
            else:
                sql = f"SELECT *, ({sql_expr}) AS {qname} FROM ({sql}) q"
                known.add(name)
        out = con.execute(sql).df()
    finally:
        con.unregister("_calc_in")
    return {"out": out}, {}


def filter_preview(pid, wid, node_id, config: dict) -> dict:
    """Dry-run d'un Filtre (semi/anti-jointure) contre ses entrées actuelles, sans
    matérialisation. Renvoie les comptes (total entrées, gardées, exclues) plus un
    statut explicite pour les configs incomplètes (paires vides, colonnes absentes,
    référence non branchée) — l'utilisateur voit *avant* d'exécuter combien de
    lignes vont survivre au filtre. Inspiré de `route_preview` pour Validation."""
    wf = storage.get_workflow(pid, wid)
    if not wf:
        raise ValueError("Workflow introuvable")
    node = next((n for n in wf.get("nodes", []) if n["id"] == node_id), None)
    if not node:
        raise ValueError("Bloc introuvable")
    d = config or {}
    ins = _node_inputs(pid, wid, node, wf.get("edges", []))
    main = next((df for a, df in ins if a == "in"), None)
    ref = next((df for a, df in ins if a == "ref"), None)
    if main is None:
        return {"status": "no_main", "message": "branchez l'entrée principale « données »"}
    total_main = int(len(main))
    if ref is None:
        return {
            "status": "no_ref",
            "message": "branchez l'entrée de référence « réf »",
            "total_main": total_main,
        }
    pairs = [p for p in (d.get("pairs") or []) if p.get("column") and p.get("ref_column")]
    if not pairs and d.get("column") and d.get("ref_column"):
        pairs = [{"column": d["column"], "ref_column": d["ref_column"]}]
    if not pairs:
        return {
            "status": "no_pairs",
            "message": "choisissez au moins une paire de colonnes à comparer",
            "total_main": total_main,
            "total_ref": int(len(ref)),
        }
    main_cols = [p["column"] for p in pairs]
    ref_cols = [p["ref_column"] for p in pairs]
    for c in main_cols:
        if c not in main.columns:
            return {
                "status": "bad_column",
                "message": f"colonne « {c} » absente des données",
                "total_main": total_main,
                "total_ref": int(len(ref)),
            }
    for c in ref_cols:
        if c not in ref.columns:
            return {
                "status": "bad_column",
                "message": f"colonne « {c} » absente de la référence",
                "total_main": total_main,
                "total_ref": int(len(ref)),
            }
    ci = bool(d.get("case_insensitive"))

    def _row_keys(df: pd.DataFrame, cols: list[str]) -> pd.Series:
        parts = [df[c].astype("string").fillna("") for c in cols]
        if ci:
            parts = [p.str.lower() for p in parts]
        return pd.Series(list(zip(*parts)), index=df.index)

    ref_keep_mask = pd.Series(True, index=ref.index)
    for c in ref_cols:
        ref_keep_mask &= ref[c].notna()
    rset = set(_row_keys(ref[ref_keep_mask], ref_cols).tolist())
    main_keep_mask = pd.Series(True, index=main.index)
    for c in main_cols:
        main_keep_mask &= main[c].notna()
    main_keys = _row_keys(main, main_cols)
    present = main_keep_mask & main_keys.isin(rset)
    matches = int(present.sum())
    null_keys = int((~main_keep_mask).sum())
    keep_mode = d.get("mode", "keep") != "exclude"
    kept = matches if keep_mode else (total_main - matches)
    excluded = total_main - kept
    return {
        "status": "ok",
        "mode": "keep" if keep_mode else "exclude",
        "total_main": total_main,
        "total_ref": int(len(ref)),
        "matches": matches,
        "kept": kept,
        "excluded": excluded,
        "null_keys": null_keys,
    }


def _run_filter(con, pid, node, ins):
    """Membership filter (semi-join / anti-join). Keeps — or excludes — the rows
    of the main input whose values in N chosen columns ALL appear together in the
    corresponding columns of the reference input (composite-key membership). The
    reference table is only consulted, never merged."""
    d = node["data"]
    main = next((df for a, df in ins if a == "in"), None)
    ref = next((df for a, df in ins if a == "ref"), None)
    if main is None:
        raise ValueError("entrée principale « données » non connectée")
    if ref is None:
        raise ValueError(
            "entrée de référence « réf » non connectée — branchez le tableau dont on lit les valeurs"
        )
    # New multi-pair model (`pairs: [{column, ref_column}, ...]`) with a fallback
    # to the legacy single-pair fields so old workflows keep running.
    pairs = [p for p in (d.get("pairs") or []) if p.get("column") and p.get("ref_column")]
    if not pairs and d.get("column") and d.get("ref_column"):
        pairs = [{"column": d["column"], "ref_column": d["ref_column"]}]
    if not pairs:
        raise ValueError("choisissez au moins une paire de colonnes à comparer")
    main_cols = [p["column"] for p in pairs]
    ref_cols = [p["ref_column"] for p in pairs]
    for c in main_cols:
        if c not in main.columns:
            raise ValueError(f"colonne « {c} » absente des données")
    for c in ref_cols:
        if c not in ref.columns:
            raise ValueError(f"colonne « {c} » absente de la référence")

    ci = bool(d.get("case_insensitive"))

    def _row_keys(df: pd.DataFrame, cols: list[str]) -> pd.Series:
        # one tuple per row, NaN → ''. With CI on, lowercase each component so
        # the comparison ignores case across the whole composite key.
        parts = [df[c].astype("string").fillna("") for c in cols]
        if ci:
            parts = [p.str.lower() for p in parts]
        return pd.Series(list(zip(*parts)), index=df.index)

    # A ref row with a NaN on any compared column has an undefined identity for
    # this composite key — drop it from the lookup set (mirrors the legacy
    # single-column dropna()).
    ref_keep_mask = pd.Series(True, index=ref.index)
    for c in ref_cols:
        ref_keep_mask &= ref[c].notna()
    rset = set(_row_keys(ref[ref_keep_mask], ref_cols).tolist())

    # Same on the main side: a row with NaN in any compared column can't match.
    main_keep_mask = pd.Series(True, index=main.index)
    for c in main_cols:
        main_keep_mask &= main[c].notna()
    main_keys = _row_keys(main, main_cols)
    present = main_keep_mask & main_keys.isin(rset)

    keep = d.get("mode", "keep") != "exclude"  # keep = semi-join, exclude = anti-join
    out = main[present if keep else ~present].reset_index(drop=True)
    return {"out": out}, {}


def _run_cols(con, pid, node, ins):
    """Reorder / drop / rename columns. The config holds an ordered list of
    {name, keep, rename}; listed columns come out in that order, dropped ones are
    removed, and any column that appeared upstream after the list was built is
    appended unchanged so nothing is silently lost."""
    d = node["data"]
    df = _single(ins)
    items = d.get("columns") or []
    pairs, seen = [], set()  # (source_name, output_name)
    for it in items:
        name = it.get("name")
        if not name or name not in df.columns or name in seen:
            continue
        seen.add(name)
        if it.get("keep") is False:
            continue
        pairs.append((name, (it.get("rename") or "").strip() or name))
    for c in df.columns:  # columns added upstream since: keep as-is
        if c not in seen:
            pairs.append((c, c))
    if not pairs:
        raise ValueError("aucune colonne conservée — gardez-en au moins une")
    out_names = [o for _, o in pairs]
    dups = {n for n in out_names if out_names.count(n) > 1}
    if dups:
        raise ValueError("noms de colonnes en double après renommage : " + ", ".join(sorted(dups)))
    out = df[[s for s, _ in pairs]].copy()
    out.columns = out_names
    return {"out": out}, {}


# Analyse block — configurable breakdowns of a column, inspired by the Validation
# filters. Each analysis derives a *category key* per row, then counts occurrences.
_ANALYSIS_TITLE = {
    "values": "Valeurs de {col}",
    "prefix": "Préfixes de {col}",
    "suffix": "Suffixes de {col}",
    "length": "Longueurs de {col}",
    "rule": "Respect d'une règle — {col}",
    "mask": "Respect d'un format — {col}",
}


def _bucket_label(k) -> str:
    if k is None or (isinstance(k, str) and k == ""):
        return "(vide)"
    if isinstance(k, float) and float(k).is_integer():
        return str(int(k))
    return str(k)


def _analysis_keys(s: pd.Series, spec: dict, kind: str) -> pd.Series:
    """Per-row category key for an analysis (a pandas Series; None = excluded)."""
    cs = bool(spec.get("case_sensitive"))
    if kind == "length":
        return s.map(lambda v: None if _is_null(v) else len(str(v)))
    if kind in ("prefix", "suffix"):
        by = spec.get("by") or "sep"
        sep = spec.get("sep") or ""
        n = int(spec.get("n") or 3)

        def key(v):
            if _is_null(v):
                return None
            t = str(v)
            if by == "chars":
                return t[:n] if kind == "prefix" else (t[-n:] if n else "")
            if not sep:
                return t
            return t.split(sep, 1)[0] if kind == "prefix" else t.rsplit(sep, 1)[-1]

        return s.map(key)
    if kind in ("rule", "mask"):
        if kind == "mask":
            pat = _mask_pattern(spec.get("segments") or [], cs)
            flags = 0 if cs else re.IGNORECASE
            ok = (lambda t: bool(re.fullmatch(pat, t, flags))) if pat else (lambda t: t == "")
        else:
            rule = spec.get("rule") or {}
            ok = lambda t: _rule_ok(t, rule, cs)  # noqa: E731
        return s.map(lambda v: "Respecté" if ok("" if _is_null(v) else str(v)) else "Non respecté")
    return s.where(s.notna(), None)  # values


def _compute_analysis(df: pd.DataFrame, spec: dict, col: str, type_str: str, total: int) -> dict:
    kind = spec.get("kind") or "values"
    s = df[col]
    nulls = int(s.isna().sum())
    keys = _analysis_keys(s, spec, kind)
    vc = keys.dropna().value_counts()
    top = max(1, int(spec.get("top") or 12))
    head = vc.head(top)
    buckets = [{"key": _bucket_label(k), "count": int(c)} for k, c in head.items()]
    shown = int(head.sum())
    other = max(0, int(len(s) - nulls) - shown)
    title = (spec.get("title") or "").strip() or _ANALYSIS_TITLE.get(kind, "{col}").format(col=col)
    return {
        "title": title,
        "column": col,
        "type": type_str,
        "kind": kind,
        "chart": spec.get("chart") or "bar",
        "total": int(total),
        "nulls": nulls,
        "null_pct": round(nulls / total * 100, 1) if total else 0.0,
        "distinct": int(keys.dropna().nunique()),
        "buckets": buckets,
        "other": other,
    }


def _keys_size_bucket(n: int) -> str:
    if n <= 3:
        return str(n)
    if n <= 5:
        return "4–5"
    if n <= 10:
        return "6–10"
    return "11+"


_KEYS_BUCKET_ORDER = ["1", "2", "3", "4–5", "6–10", "11+"]


def _compute_keys_analysis(df: pd.DataFrame, spec: dict, types: dict, total: int) -> dict:
    """Multi-key / cardinality analysis. Groups rows by a key (one or several
    columns) and profiles the *groups*: duplicate/uniqueness stats (is this a
    candidate key?), the group-size distribution, the biggest groups with sample
    rows, and per-column functional-dependency consistency — for each non-key
    column, is it constant within a key (key → column) or does it vary?"""
    key_cols = [c for c in (spec.get("key") or []) if c in types]
    if not key_cols and spec.get("column") in types:
        key_cols = [spec["column"]]
    title = (spec.get("title") or "").strip() or ("Clés multiples — " + ", ".join(key_cols))
    if not key_cols:
        return {
            "kind": "keys",
            "title": title,
            "key": [],
            "total": int(total),
            "error": "Aucune colonne clé valide — choisissez au moins une colonne.",
        }

    top = max(1, int(spec.get("top") or 12))
    sample_rows = 20
    g = df.groupby(key_cols, dropna=False, sort=False)
    sizes = g.size()
    sz = sizes.to_numpy()
    distinct_keys = int(len(sizes))
    dup_mask = sizes > 1
    dup_keys = int(dup_mask.sum())
    rows_in_dups = int(sizes[dup_mask].sum())
    null_keys = int(df[key_cols].isna().any(axis=1).sum())

    # group frames keyed by a normalized tuple (pandas yields scalar keys for a
    # single grouper, tuples for several — normalize so lookups are uniform).
    _astuple = lambda k: k if isinstance(k, tuple) else (k,)  # noqa: E731
    frames = {_astuple(name): sub for name, sub in g}

    counts: dict[str, int] = {}
    for n in sz:
        b = _keys_size_bucket(int(n))
        counts[b] = counts.get(b, 0) + 1
    distribution = [
        {"key": b, "count": int(counts[b])} for b in _KEYS_BUCKET_ORDER if counts.get(b)
    ]

    def _groups_from(sorted_sizes):
        out = []
        for kv, n in sorted_sizes.items():
            kvt = _astuple(kv)
            out.append(
                {
                    "key": {c: _json_safe(v) for c, v in zip(key_cols, kvt)},
                    "size": int(n),
                    "rows": _df_records(frames[kvt].head(sample_rows)),
                }
            )
        return out

    top_groups = _groups_from(sizes.sort_values(ascending=False, kind="stable").head(top))
    small_groups = _groups_from(sizes.sort_values(ascending=True, kind="stable").head(top))

    consistency = []
    for c in (col for col in df.columns if col not in key_cols):
        nun = g[c].nunique(dropna=False)
        varying = int((nun > 1).sum())
        entry = {"column": c, "varying_groups": varying, "constant": varying == 0}
        if varying:
            kvt = _astuple(nun[nun > 1].index[0])
            vals = frames[kvt][c].dropna().unique().tolist()[:6]
            entry["example"] = {
                "key": {kc: _json_safe(v) for kc, v in zip(key_cols, kvt)},
                "values": [_json_safe(v) for v in vals],
            }
        consistency.append(entry)
    consistency.sort(key=lambda e: (e["constant"], -e["varying_groups"]))

    return {
        "kind": "keys",
        "title": title,
        "key": key_cols,
        "total": int(total),
        "distinct_keys": distinct_keys,
        "dup_keys": dup_keys,
        "singletons": int((sizes == 1).sum()),
        "rows_in_dups": rows_in_dups,
        "null_keys": null_keys,
        "unique": bool(distinct_keys == total and total > 0),
        "uniqueness_pct": round(distinct_keys / total * 100, 1) if total else 0.0,
        "group_min": int(sz.min()) if len(sz) else 0,
        "group_max": int(sz.max()) if len(sz) else 0,
        "group_avg": round(float(sz.mean()), 2) if len(sz) else 0.0,
        "group_median": float(np.median(sz)) if len(sz) else 0.0,
        "distribution": distribution,
        "top_groups": top_groups,
        "small_groups": small_groups,
        "consistency": consistency,
    }


def _run_report(con, pid, node, ins):
    """Analyse block: terminal (no downstream output). Passes its input through
    (materialized so it stays previewable) and computes the configured analyses —
    column breakdowns with a chosen chart — stored on the meta for the block view
    and the documentation export."""
    d = node["data"]
    df = _single(ins)
    con.register("_rep_in", df)
    try:
        types = {c["name"]: c["type"] for c in _describe(con, "_rep_in")}
        total = con.execute("SELECT count(*) FROM _rep_in").fetchone()[0]
    finally:
        con.unregister("_rep_in")
    specs = d.get("analyses")
    if not specs:  # back-compat: one 'values' analysis per chosen column
        specs = [
            {"column": c, "kind": "values", "chart": "bar"}
            for c in (d.get("columns") or list(types))
            if c in types
        ]
    analyses = []
    for spec in specs:
        if (spec.get("kind") or "values") == "keys":
            analyses.append(_compute_keys_analysis(df, spec, types, total))
            continue
        col = spec.get("column")
        if col and col in types:
            analyses.append(_compute_analysis(df, spec, col, types[col], total))
    report = {"note": (d.get("note") or "").strip(), "row_count": int(total), "analyses": analyses}
    return {"out": df}, {"report": report}


def _run_union(con, pid, node, ins):
    d = node["data"]
    if not ins:
        raise ValueError("aucune entrée connectée")
    names = []
    for i, (_, df) in enumerate(ins):
        nm = f"_u{i}"
        con.register(nm, df)
        names.append(nm)
    by_name = d.get("by_name", True) is not False
    try:
        if len(names) == 1:
            sql = f"SELECT * FROM {names[0]}"
        else:
            joiner = " UNION ALL BY NAME " if by_name else " UNION ALL "
            sql = joiner.join(f"SELECT * FROM {nm}" for nm in names)
        if d.get("distinct"):
            sql = f"SELECT DISTINCT * FROM ({sql})"
        out = con.execute(sql).df()
    finally:
        for nm in names:
            con.unregister(nm)
    return {"out": out}, {}


def _sanitize_filename(name: str, default: str) -> str:
    return re.sub(r"[\\/:*?\"<>|]", "_", str(name or "").strip()) or default


def _sanitize_sheet_name(name: str) -> str:
    """Excel sheet-name rules: max 31 chars, none of []:*?/\\, never blank."""
    s = re.sub(r"[\\/:*?\[\]]", "_", str(name or "")).strip()
    return s[:31] or "Feuille"


def _run_export(
    con, pid, node, ins, export_sink=None, workflow_name=None, force_enabled=False, produced=None
):
    d = node["data"]
    if d.get("enabled") is False and not force_enabled:  # disabled export: don't touch the file
        return {}, {"skipped": True, "row_count": 0}
    df = _single(ins)
    n = int(len(df))
    if n == 0:  # empty result: write nothing at all
        return {}, {"skipped_empty": True, "row_count": 0}
    fn = _sanitize_filename(d.get("filename"), "resultat")
    out_dir = storage.workflow_export_dir(pid, workflow_name or "Workflow")
    out_dir.mkdir(parents=True, exist_ok=True)

    # multi-sheet mode: contribute one sheet to a single workbook named after the
    # workflow. The actual write is deferred to the end of the run so every sheet
    # lands in the same file (see _write_workbook in iter_run_workflow).
    if d.get("to_workbook") and export_sink is not None:
        wb = _sanitize_filename(workflow_name, "Workflow")
        wb_path = out_dir / f"{wb}.xlsx"
        sheet = _sanitize_sheet_name(fn)
        export_sink.setdefault(str(wb_path), []).append((sheet, df))
        if produced is not None:
            produced.add(wb_path.name)
        return {}, {"exported": f"{wb}.xlsx", "sheet": sheet, "row_count": n, "deferred": True}

    fmt = (d.get("format") or "xlsx").lower()
    if fmt == "csv":
        out_path = out_dir / f"{fn}.csv"
        _write_csv(df, out_path, d.get("csv") or {})
    else:
        out_path = out_dir / f"{fn}.xlsx"
        _check_excel_limits(df)  # garde-fou A.7 : on échoue AVANT d'écrire
        df.to_excel(out_path, index=False)
    if produced is not None:
        produced.add(out_path.name)
    return {}, {"exported": out_path.name, "row_count": n}


# Limites « hard » d'Excel — un export qui les dépasse échoue *après* tout le calcul.
# Il vaut mieux le détecter avant l'écriture, avec une suggestion (cf. todo A.7).
EXCEL_MAX_ROWS = 1_048_576  # 2**20
EXCEL_MAX_COLS = 16_384  # 2**14


def _check_excel_limits(df: pd.DataFrame) -> None:
    n = len(df)
    # +1 pour la ligne d'en-tête
    if n + 1 > EXCEL_MAX_ROWS:
        raise ValueError(
            f"Excel limite à {EXCEL_MAX_ROWS:,} lignes (vous en auriez {n:,}). "
            "Suggestion : exportez en CSV (format paramétrable), ou découpez en "
            "plusieurs feuilles / fichiers."
        )
    nc = len(df.columns)
    if nc > EXCEL_MAX_COLS:
        raise ValueError(
            f"Excel limite à {EXCEL_MAX_COLS:,} colonnes (vous en auriez {nc:,}). "
            "Suggestion : utilisez le bloc Colonnes pour réduire, ou transposez."
        )


def _write_csv(df: pd.DataFrame, path, opts: dict) -> None:
    """Export CSV paramétrable (cf. todo A.4).

    Preset par défaut : « Excel FR » — séparateur `;`, décimale `,`, encodage
    UTF-8 BOM (`utf-8-sig`), fin de ligne `\\r\\n`. Excel FR autodétecte alors
    correctement les accents et le format des nombres.
    """
    sep = opts.get("sep") or ";"
    decimal = opts.get("decimal") or ","
    bom = opts.get("bom", True)
    eol = opts.get("eol") or "\r\n"
    encoding = "utf-8-sig" if bom else "utf-8"
    df.to_csv(
        path,
        index=False,
        sep=sep,
        decimal=decimal,
        encoding=encoding,
        lineterminator=eol,
    )


def _write_workbook(path, sheets, fresh=True):
    """Write the gathered (sheet_name, DataFrame) pairs into one .xlsx workbook.
    A full run (fresh) rewrites the file from scratch; a partial run appends/
    replaces only its own sheets, leaving the others intact. Duplicate sheet
    names are disambiguated (Excel forbids them)."""
    import pathlib

    sheets = [(s, df) for s, df in sheets if df is not None and len(df)]
    if not sheets:  # nothing to write -> no empty file
        return
    seen, final = set(), []
    for name, df in sheets:
        base = name or "Feuille"
        nm, k = base, 2
        while nm.lower() in seen:
            sfx = f"_{k}"
            nm = base[: 31 - len(sfx)] + sfx
            k += 1
        seen.add(nm.lower())
        final.append((nm, df))
    # Garde-fou A.7 sur chaque feuille AVANT d'ouvrir le writer (sinon on
    # écrit partiellement puis on plante openpyxl, laissant un fichier corrompu).
    for _, df in final:
        _check_excel_limits(df)
    p = pathlib.Path(path)
    if fresh or not p.exists():
        writer = pd.ExcelWriter(p, engine="openpyxl", mode="w")
    else:
        writer = pd.ExcelWriter(p, engine="openpyxl", mode="a", if_sheet_exists="replace")
    with writer:
        for nm, df in final:
            df.to_excel(writer, sheet_name=nm, index=False)


# export_filenames was used to tell exports apart from sources back when both
# lived in projects/<pid>/files/. Exports now live in projects/<pid>/exports/<wf>/,
# tagged explicitly at listing time — no inference from filenames required.


_RUNNERS = {
    "source": _run_source,
    "sql": _run_sql,
    "dedup": _run_dedup,
    "validate": _run_validate,
    "pivot": _run_pivot,
    "clean": _run_clean,
    "calc": _run_calc,
    "union": _run_union,
    "export": _run_export,
    "filter": _run_filter,
    "cols": _run_cols,
    "report": _run_report,
}


# --------------------------------------------------------------------------- #
# Validation logic (also used by the live tester endpoint)
# --------------------------------------------------------------------------- #
_CLASS_RE = {
    "letter": "[A-Za-z]",
    "upper": "[A-Z]",
    "lower": "[a-z]",
    "digit": "[0-9]",
    "alnum": "[A-Za-z0-9]",
    "any": ".",
}


def _seg_pattern(seg) -> str:
    """Regex for one mask segment."""
    t = seg.get("type") or "any"
    if t == "literal":
        return re.escape(seg.get("value") or "")
    base = "[" + re.escape(seg.get("value") or "") + "]" if t == "set" else _CLASS_RE.get(t, ".")
    ln, lo, hi = seg.get("length"), seg.get("min"), seg.get("max")
    # % format lisible pour des quantifieurs regex `{n}` / `{a,b}` ; en f-string,
    # les `{{` / `}}` rendent l'intention illisible.
    if ln not in (None, ""):
        return base + "{%d}" % int(ln)  # noqa: UP031
    if lo not in (None, "") or hi not in (None, ""):
        return base + "{%s,%s}" % (  # noqa: UP031
            int(lo) if lo not in (None, "") else 0,
            int(hi) if hi not in (None, "") else "",
        )
    return base + "{1}"


def _mask_pattern(segments: list, case_sensitive: bool = False) -> str:
    return "".join(_seg_pattern(s) for s in segments or [])


def _segment_label(seg) -> str:
    t = seg.get("type") or "any"
    if t == "literal":
        return f"« {seg.get('value') or ''} »"
    names = {
        "letter": "lettres",
        "upper": "MAJ",
        "lower": "min",
        "digit": "chiffres",
        "alnum": "alphanum.",
        "set": f"[{seg.get('value') or ''}]",
        "any": "tout",
    }
    base = names.get(t, t)
    ln, lo, hi = seg.get("length"), seg.get("min"), seg.get("max")
    if ln not in (None, ""):
        return f"{base}×{int(ln)}"
    if lo not in (None, "") or hi not in (None, ""):
        return f"{base}×{lo or 0}–{hi if hi not in (None, '') else '∞'}"
    return base


def _mask_steps(s: str, segments: list, cs: bool) -> list[dict]:
    """Per-segment status for one sample: the first segment whose cumulative
    prefix no longer matches is the blocker ('fail'); earlier ones are 'ok',
    later ones 'skip'. A trailing extra (string longer than the mask) is flagged."""
    flags = 0 if cs else re.IGNORECASE
    segs = segments or []
    pats = [_seg_pattern(x) for x in segs]
    labels = [_segment_label(x) for x in segs]
    n = len(pats)
    if n == 0:
        return []
    blocker = n
    for k in range(n):
        if not re.match("".join(pats[: k + 1]), s, flags):
            blocker = k
            break
    steps = [
        {
            "label": labels[i],
            "status": "ok" if i < blocker else ("fail" if i == blocker else "skip"),
        }
        for i in range(n)
    ]
    if blocker == n and not re.fullmatch("".join(pats), s, flags):
        steps.append({"label": "caractères en trop", "status": "fail"})
    return steps


def _char_in_class(ch: str, cls: str, cs: bool) -> bool:
    if not cs and cls in ("upper", "lower"):
        return ch.isalpha()
    return {
        "letter": ch.isalpha(),
        "upper": ch.isupper(),
        "lower": ch.islower(),
        "digit": ch.isdigit(),
        "alnum": ch.isalnum(),
    }.get(cls, True)


def _rule_label(r: dict) -> str:
    if r.get("label"):
        return r["label"]
    test = r.get("test")
    # Mode multi-valeur : on affiche la liste pour que les libellés des règles
    # restent lisibles (« doit commencer par l'une de : 00, 01, 02… »).
    mv = _multi_values(r)
    if mv is not None:
        preview = ", ".join(str(x) for x in mv[:6])
        if len(mv) > 6:
            preview += f"… (+{len(mv) - 6})"
        multi_labels = {
            "starts_with": f"doit commencer par l'une de : {preview}",
            "ends_with": f"doit finir par l'une de : {preview}",
            "contains": f"doit contenir l'une de : {preview}",
            "not_contains": f"ne doit contenir aucune de : {preview}",
            "equals": f"doit être égal à l'une de : {preview}",
            "not_equals": f"ne doit être égal à aucune de : {preview}",
            "regex": f"doit correspondre à l'une de : {preview}",
            "regex_full": f"doit correspondre à l'une de : {preview}",
            "char_equals": f"le caractère {r.get('position', 1)} doit être l'un de : {preview}",
            "substr_equals": f"les caractères {r.get('start', 1)}… doivent être l'un de : {preview}",
        }
        if test in multi_labels:
            return multi_labels[test]
    v = r.get("value", "")
    labels = {
        "starts_with": f"doit commencer par « {v} »",
        "ends_with": f"doit finir par « {v} »",
        "contains": f"doit contenir « {v} »",
        "not_contains": f"ne doit pas contenir « {v} »",
        "equals": f"doit être égal à « {v} »",
        "not_equals": f"doit être différent de « {v} »",
        "regex": f"doit correspondre à {v}",
        "regex_full": f"doit correspondre à {v}",
        "length_eq": f"longueur doit être {v}",
        "length_min": f"longueur ≥ {v}",
        "length_max": f"longueur ≤ {v}",
        "char_class": f"le caractère {r.get('position', 1)} est invalide",
        "char_equals": f"le caractère {r.get('position', 1)} doit être « {v} »",
        "substr_equals": f"les caractères {r.get('start', 1)}… doivent être « {v} »",
        "substr_class": f"les caractères {r.get('start', 1)}… sont invalides",
        "is_in": "valeur hors liste autorisée",
        "is_empty": "doit être vide",
        "not_empty": "ne doit pas être vide",
        "is_numeric": "doit être numérique",
    }
    return labels.get(test, test or "règle non respectée")


# Tests that have a clean opposite — so a negated rule reads as its positive twin
# instead of an unreadable double negative ("NON ne doit pas contenir" -> "doit
# contenir").
_RULE_OPPOSITE = {
    "contains": "not_contains",
    "not_contains": "contains",
    "equals": "not_equals",
    "not_equals": "equals",
    "is_empty": "not_empty",
    "not_empty": "is_empty",
}


def _signed_label(r: dict) -> str:
    """Human label for a (possibly negated) rule, folding the NON into the wording
    so it never reads as a double negative."""
    if not r.get("negate"):
        return _rule_label(r)
    opp = _RULE_OPPOSITE.get(r.get("test"))
    if opp:
        return _rule_label({**r, "test": opp})
    base = _rule_label(r)  # no clean opposite: « doit … » -> « ne doit pas … »
    if "doit " in base:
        return base.replace("doit ", "ne doit pas ", 1)
    return "NON — " + base


def _rule_ok(s: str, r: dict, cs: bool) -> bool:
    # Multi-valeur (cf. _rule_mask) — même sémantique côté scalaire pour rester
    # cohérent quand un test non vectorisable est utilisé. La négation `r.negate`
    # est appliquée par le caller (testers / _rule_mask), pas ici : on renvoie
    # le résultat brut du test.
    mv = _multi_values(r)
    if mv is not None:
        is_neg = r.get("test") in _NEGATION_OF
        pos_test = _NEGATION_OF.get(r.get("test"), r.get("test"))
        sub_base = {k: v for k, v in r.items() if k not in ("values", "negate")}
        sub_base["test"] = pos_test
        ok = any(_rule_ok(s, {**sub_base, "value": v}, cs) for v in mv)
        return (not ok) if is_neg else ok

    test = r.get("test")
    v = r.get("value")
    flags = 0 if cs else re.IGNORECASE

    def norm(x):
        return x if cs else str(x).lower()

    ss = s if cs else s.lower()
    vv = norm(v) if v is not None else ""
    try:
        if test == "starts_with":
            return ss.startswith(vv)
        if test == "ends_with":
            return ss.endswith(vv)
        if test == "contains":
            return vv in ss
        if test == "not_contains":
            return vv not in ss
        if test == "equals":
            return ss == vv
        if test == "not_equals":
            return ss != vv
        if test == "regex":
            return re.search(str(v), s, flags) is not None
        if test == "regex_full":
            return re.fullmatch(str(v), s, flags) is not None
        if test == "length_eq":
            return len(s) == int(v)
        if test == "length_min":
            return len(s) >= int(v)
        if test == "length_max":
            return len(s) <= int(v)
        if test == "char_class":
            p = int(r.get("position") or 1)
            return 0 < p <= len(s) and _char_in_class(s[p - 1], r.get("charclass", "letter"), cs)
        if test == "char_equals":
            p = int(r.get("position") or 1)
            return 0 < p <= len(s) and norm(s[p - 1]) == vv
        if test == "substr_equals":
            st = int(r.get("start") or 1)
            ln = int(r.get("length") or 1)
            return norm(s[st - 1 : st - 1 + ln]) == vv
        if test == "substr_class":
            st = int(r.get("start") or 1)
            ln = int(r.get("length") or 1)
            seg = s[st - 1 : st - 1 + ln]
            return len(seg) == ln and all(
                _char_in_class(c, r.get("charclass", "letter"), cs) for c in seg
            )
        if test == "is_in":
            items = [norm(x.strip()) for x in str(v).split(",") if x.strip() != ""]
            return ss in items
        if test == "is_empty":
            return s == ""
        if test == "not_empty":
            return s != ""
        if test == "is_numeric":
            float(str(s).replace(",", "."))
            return True
        if test in ("num_eq", "num_ne", "num_gt", "num_ge", "num_lt", "num_le"):
            # The per-row tester has no other column in scope; "vs column" mode
            # is a no-op here (treated as a passing rule so it doesn't poison
            # the live tester). Vectorized routing handles the real comparison.
            if r.get("via") == "column":
                return True
            a = float(str(s).replace(",", "."))
            b = float(str(v).replace(",", "."))
            return {
                "num_eq": a == b,
                "num_ne": a != b,
                "num_gt": a > b,
                "num_ge": a >= b,
                "num_lt": a < b,
                "num_le": a <= b,
            }[test]
    except (ValueError, TypeError):
        return False
    return True


def _validate_one(value, cfg: dict) -> tuple[bool, str]:
    cs = bool(cfg.get("case_sensitive"))
    s = "" if _is_null(value) else str(value)
    if cfg.get("mode") == "mask":
        pat = _mask_pattern(cfg.get("segments") or [], cs)
        flags = 0 if cs else re.IGNORECASE
        ok = bool(re.fullmatch(pat, s, flags)) if pat else (s == "")
        return ok, ("" if ok else "ne correspond pas au format attendu")
    rules = cfg.get("rules") or []
    if not rules:
        return True, ""
    combine = cfg.get("combine", "all")
    fails = [r for r in rules if not _rule_ok(s, r, cs)]
    if combine == "any":
        ok = len(fails) < len(rules)
        return ok, ("" if ok else "aucune règle satisfaite")
    ok = not fails
    return ok, ("" if ok else _rule_label(fails[0]))


def _condition_eval_str(s: str, cond: dict, cs: bool) -> tuple[bool, str, list]:
    """Evaluate a named condition against a single string sample, returning
    (valid, motif, per-step statuses) for the live tester."""
    cond = cond or {}
    if cond.get("kind") == "mask":
        pat = _mask_pattern(cond.get("segments") or [], cs)
        flags = 0 if cs else re.IGNORECASE
        ok = bool(re.fullmatch(pat, s, flags)) if pat else (s == "")
        steps = _mask_steps(s, cond.get("segments") or [], cs)
        return ok, ("" if ok else "ne correspond pas au format attendu"), steps
    groups = cond.get("groups") or []
    if not groups:
        return True, "", []
    steps, group_ok, first_fail = [], [], None
    for g in groups:
        rules = g.get("rules") or []
        oks = []
        for r in rules:
            rok = _rule_ok(s, r, cs)
            if r.get("negate"):
                rok = not rok
            oks.append(rok)
            label = _signed_label(r)
            steps.append({"label": label, "status": "ok" if rok else "fail"})
            if not rok and first_fail is None:
                first_fail = label
        group_ok.append(all(oks) if rules else False)
    ok = any(group_ok)
    motif = "" if ok else (first_fail or "aucune règle satisfaite")
    return ok, motif, steps


def validate_test(config: dict, samples: list) -> dict:
    cfg = config or {}
    cs = bool(cfg.get("case_sensitive"))
    try:
        results = []
        if cfg.get("kind") in ("rules", "mask"):  # new: a single named condition
            for x in samples or []:
                s = "" if _is_null(x) else str(x)
                ok, motif, steps = _condition_eval_str(s, cfg, cs)
                results.append(
                    {
                        "valid": ok,
                        "motif": "OK" if ok else (motif or "non conforme"),
                        "steps": steps,
                    }
                )
            return {"ok": True, "results": results}
        is_mask = cfg.get("mode") == "mask"  # legacy: rules/mask conformity block
        for x in samples or []:
            s = "" if _is_null(x) else str(x)
            ok, motif = _validate_one(x, cfg)
            if is_mask:
                steps = _mask_steps(s, cfg.get("segments") or [], cs)
            else:
                steps = [
                    {"label": _rule_label(r), "status": "ok" if _rule_ok(s, r, cs) else "fail"}
                    for r in (cfg.get("rules") or [])
                ]
            results.append(
                {"valid": ok, "motif": "OK" if ok else (motif or "non conforme"), "steps": steps}
            )
        return {"ok": True, "results": results}
    except Exception as e:  # noqa: BLE001 - surfaced to the UI tester
        return {"ok": False, "error": str(e)}


# --------------------------------------------------------------------------- #
# Materialization
# --------------------------------------------------------------------------- #
def _empty_parquet_rows(path):
    """A parquet written from a frame with NO columns — an 'empty' routed output
    (e.g. an else/route branch fed by a 0-column input). DuckDB refuses these
    ('Need at least one non-root column'); pandas reads them. Returns the row count
    for such a file, else None (a normal parquet -> keep the DuckDB path)."""
    try:
        import pyarrow.parquet as pq

        md = pq.read_metadata(path)
        if md.num_columns == 0:
            return int(md.num_rows)
    except Exception:
        pass
    return None


def _meta_from_parquet(path) -> dict:
    n = _empty_parquet_rows(path)
    if n is not None:  # 0-column 'empty' output: skip DuckDB
        return {"columns": [], "row_count": n, "sample": [], "stats": []}
    con = duckdb.connect()
    try:
        rel = f"read_parquet({_lit(path)})"
        columns = _describe(con, rel)
        row_count = con.execute(f"SELECT count(*) FROM {rel}").fetchone()[0]
        sample = _df_records(con.execute(f"SELECT * FROM {rel} LIMIT 50").df())
        try:
            stats = _df_records(con.execute(f"SUMMARIZE SELECT * FROM {rel}").df())
        except Exception:
            stats = []
        return {"columns": columns, "row_count": int(row_count), "sample": sample, "stats": stats}
    finally:
        con.close()


def _write_output(pid, wid, node, handle, df, extra) -> dict:
    storage.data_dir(pid, wid).mkdir(parents=True, exist_ok=True)
    path = storage.node_parquet(pid, wid, node["id"], handle)
    df.to_parquet(path, index=False)
    meta = {"block_type": node["type"], "label": node["data"].get("label", node["id"])}
    meta.update(_meta_from_parquet(path))
    meta.update(extra)
    storage.write_node_meta(pid, wid, node["id"], meta, handle)
    return meta


def _reuse_result(pid, wid, node, handles, ntype, locked, signature) -> dict:
    """Build a 'cached' run result from the stored meta of an existing output."""
    outputs, columns, extra = {}, [], {}
    for i, h in enumerate(handles):
        m = storage.read_node_meta(pid, wid, node["id"], h) or {}
        outputs[h] = m.get("row_count", 0)
        if i == 0:
            columns = m.get("columns", [])
            for k in ("compiled_sql", "clean_report", "report", "casts_report", "mixed_types"):
                if k in m:
                    extra[k] = m[k]
            for k in ("input_rows", "unrouted_rows"):
                if k in m:
                    extra[k] = m[k]
    return {
        "type": ntype,
        "locked": locked,
        "cached": True,
        "signature": signature,
        "outputs": outputs,
        "row_count": outputs.get(handles[0], 0),
        "columns": columns,
        **extra,
    }


def _execute_node(
    con,
    pid,
    wid,
    node,
    edges,
    bypass_cache=False,
    bypass_lock=False,
    export_sink=None,
    wf_name=None,
    signature=None,
    all_exports=False,
    produced=None,
    on_progress=None,
) -> dict:
    ntype = node["type"]
    handles = _output_handles(node)
    have_outputs = bool(handles) and all(
        storage.node_parquet(pid, wid, node["id"], h).exists() for h in handles
    )

    # locked: reuse the frozen result without recomputing, whatever changed.
    # A bulk "force recompute all" busts the incremental cache (bypass_cache) but
    # still honours locks; only an explicit single-node force overrides a lock.
    if node["data"].get("locked") and not bypass_lock and have_outputs:
        return _reuse_result(pid, wid, node, handles, ntype, True, signature)

    # incremental cache: reuse when the node's signature is unchanged since the
    # last run (its config and every upstream block are identical). This is what
    # makes a full run recompute only the new/changed blocks and their dependents.
    if (
        signature
        and not bypass_cache
        and have_outputs
        and node["data"].get("cache", True) is not False
    ):
        m0 = storage.read_node_meta(pid, wid, node["id"], handles[0]) or {}
        if m0.get("node_signature") == signature:
            return _reuse_result(pid, wid, node, handles, ntype, False, signature)

    ins = _node_inputs(pid, wid, node, edges)
    if ntype == "export":
        out_dfs, extra = _run_export(
            con,
            pid,
            node,
            ins,
            export_sink=export_sink,
            workflow_name=wf_name,
            force_enabled=all_exports,
            produced=produced,
        )
    else:
        if ntype == "source":
            out_dfs, extra = _run_source(con, pid, node, ins, on_progress=on_progress)
        else:
            runner = _RUNNERS.get(ntype)
            if runner is None:
                raise ValueError(f"type de bloc inconnu : {ntype}")
            out_dfs, extra = runner(con, pid, node, ins)

    if not handles:  # export (always runs; nothing to cache)
        return {
            "type": ntype,
            "locked": False,
            "cached": False,
            "signature": signature,
            "row_count": extra.get("row_count", 0),
            "outputs": {},
            "columns": [],
            **extra,
        }

    outputs, columns = {}, []
    for i, h in enumerate(handles):
        df = out_dfs.get(h)
        if df is None:
            df = pd.DataFrame()
        ex = dict(extra) if i == 0 else {}
        if i == 0 and signature:  # stamp the signature for next-run reuse
            ex["node_signature"] = signature
        meta = _write_output(pid, wid, node, h, df, ex)
        outputs[h] = meta["row_count"]
        if i == 0:
            columns = meta["columns"]
    # Now that the new outputs are written, sweep any parquet/meta from a
    # previous configuration whose handle is no longer exposed (output removed,
    # renamed-to-new-id, `else` toggled off). Otherwise an edge in the workflow
    # that still references the old handle would silently keep ingesting the
    # ghost file at the next downstream run.
    storage.prune_node_outputs(pid, wid, node["id"], handles)
    result = {
        "type": ntype,
        "locked": False,
        "cached": False,
        "signature": signature,
        "outputs": outputs,
        "row_count": outputs.get(handles[0], 0),
        "columns": columns,
    }
    for k in (
        "compiled_sql",
        "clean_report",
        "report",
        "casts_report",
        "input_rows",
        "unrouted_rows",
    ):
        if k in extra:
            result[k] = extra[k]
    return result


# E.2 — drapeaux d'annulation. Un POST sur `/cancel` ajoute (pid, wid) ici, et
# `iter_run_workflow` regarde entre chaque bloc. Simple set en mémoire : pas de
# besoin de DB tant que le runner est mono-process (cf. graine 3 cloud — quand
# on aura un broker, ce sera là qu'il faudra changer).
_cancel_requested: set[tuple[str, str]] = set()


def request_cancel(pid: str, wid: str) -> None:
    """Demande l'annulation d'un run en cours. Le runner verra le drapeau
    entre 2 blocs (granularité : un bloc en cours d'exécution ne s'arrête
    pas en plein milieu — il termine, puis le run s'arrête)."""
    _cancel_requested.add((pid, wid))


def _execute_node_with_progress(con, pid, wid, nid, node, edges, **kwargs):
    """Exécute un bloc. Pour les sources, émet des events `node_progress` via yield."""
    if node["type"] != "source":
        yield {"_kind": "result", "result": _execute_node(con, pid, wid, node, edges, **kwargs)}
        return

    q: queue.Queue = queue.Queue()

    def on_progress(rows_read, total_rows):
        q.put(
            {
                "event": "node_progress",
                "node_id": nid,
                "rows_read": int(rows_read),
                "total_rows": int(total_rows),
            }
        )

    def worker():
        try:
            res = _execute_node(con, pid, wid, node, edges, on_progress=on_progress, **kwargs)
            q.put({"_kind": "result", "result": res})
        except Exception as exc:  # noqa: BLE001
            q.put({"_kind": "error", "exc": exc})

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()
    while True:
        item = q.get()
        if item.get("_kind") == "result":
            yield item
            break
        if item.get("_kind") == "error":
            raise item["exc"]
        yield item
    thread.join()


def iter_run_workflow(pid, wid, only_node=None, force=False, all_exports=False):
    wf = storage.get_workflow(pid, wid)
    if not wf:
        yield {"event": "error", "message": "Workflow introuvable"}
        return
    nodes = {n["id"]: n for n in wf.get("nodes", [])}
    edges = wf.get("edges", [])
    wf_name = wf.get("name") or "Workflow"
    order = _topo_order(nodes, edges)
    if only_node and only_node in nodes:
        run_set = _ancestors(only_node, edges) | {only_node}
    else:
        run_set = set(nodes)
    # visual-only nodes never execute: frames (legacy name "group") and "stop"
    # caps (mark an output as closed on purpose; they consume but produce nothing)
    run_ids = [
        nid
        for nid in order
        if nid in run_set and nodes[nid]["type"] not in ("frame", "group", "stop")
    ]

    # Announce the run before the (cheap) signature pass so the UI clears its
    # "préparation…" label immediately and starts naming blocks as they run.
    yield {"event": "start", "total": len(run_ids)}
    sig_by_node = _signatures_for(pid, wid, nodes, edges, order)

    # A.10 — historique de runs : on collecte un résumé du run en cours et on
    # l'append à `runs.jsonl` à la fin. Utile pour comparer deux exécutions,
    # voir ce qui a été recalculé vs servi par le cache, repérer un bloc lent.
    import time as _time

    run_started_at = _time.time()
    run_record = {
        "started_at_ms": int(run_started_at * 1000),
        "only_node": only_node,
        "force": force,
        "ran": [],  # blocs effectivement recalculés
        "cached": [],  # blocs servis par le cache
        "errors": [],  # {node_id, label, message}
        "status": "ok",  # ok | error | aborted
    }

    # If the workflow was renamed since last run, move its export folder so the
    # on-disk layout follows the rename instead of leaving orphans behind.
    # Only acts on a FULL run (only_node=None) — a single-node run is too narrow
    # a scope to reorganize the whole project on the user's behalf.
    last = wf.get("last_exports") or {}
    target_dir = storage.workflow_export_dir(pid, wf_name)
    if only_node is None and last.get("dir") and last["dir"] != target_dir.name:
        old_dir = storage.exports_dir(pid) / last["dir"]
        if old_dir.exists() and not target_dir.exists():
            target_dir.parent.mkdir(parents=True, exist_ok=True)
            old_dir.rename(target_dir)

    con = duckdb.connect()
    export_sink = {}  # workbook_path -> [(sheet, df), …]
    produced = set()  # filenames written by this run
    cancel_key = (pid, wid)
    _cancel_requested.discard(cancel_key)  # drapeaux d'un run précédent
    try:
        for nid in run_ids:
            # E.2 — granularité d'annulation : entre 2 blocs. Un bloc en cours
            # termine son calcul (sinon on laisserait son parquet à moitié).
            if cancel_key in _cancel_requested:
                _cancel_requested.discard(cancel_key)
                run_record["status"] = "aborted"
                yield {"event": "aborted"}
                _persist_run_record(pid, wid, run_record, run_started_at)
                return
            node = nodes[nid]
            yield {
                "event": "node_start",
                "node_id": nid,
                "label": node["data"].get("label", nid),
                "type": node["type"],
            }
            single_force = force and only_node is not None and nid == only_node
            node_started = _time.time()
            try:
                result = None
                exec_kwargs = dict(
                    bypass_cache=(single_force or (force and only_node is None)),
                    bypass_lock=single_force,
                    all_exports=all_exports,
                    export_sink=export_sink,
                    wf_name=wf_name,
                    signature=sig_by_node.get(nid),
                    produced=produced,
                )
                for item in _execute_node_with_progress(
                    con, pid, wid, nid, node, edges, **exec_kwargs
                ):
                    if item.get("_kind") == "result":
                        result = item["result"]
                    else:
                        yield item
            except Exception as e:  # noqa: BLE001 - reported to the client
                label = node["data"].get("label", nid)
                run_record["errors"].append({"node_id": nid, "label": label, "message": str(e)})
                run_record["status"] = "error"
                yield {"event": "error", "node_id": nid, "message": f"{label} : {e}"}
                # On persist quand même le run partiel : pouvoir diagnostiquer un
                # échec après coup compte autant que tracer un succès.
                _persist_run_record(pid, wid, run_record, run_started_at)
                return
            elapsed_ms = int((_time.time() - node_started) * 1000)
            entry = {
                "node_id": nid,
                "label": node["data"].get("label", nid),
                "type": node["type"],
                "elapsed_ms": elapsed_ms,
                "row_count": result.get("row_count", 0),
            }
            (run_record["cached"] if result.get("cached") else run_record["ran"]).append(entry)
            yield {"event": "node_done", "node_id": nid, "result": result}
        # write the multi-sheet workbooks once every sheet has been gathered
        for wb_path, sheets in export_sink.items():
            _write_workbook(wb_path, sheets, fresh=(only_node is None))
        # Cleanup: only on a full run, only files WE previously wrote, only if
        # they're in our export dir, and only if they're not in this run's set.
        # Anything the user dropped in the folder by hand is never touched.
        if only_node is None:
            stale = set(last.get("files") or []) - produced
            for name in stale:
                p = target_dir / name
                if p.exists() and p.is_file() and p.parent == target_dir:
                    p.unlink()
        # Persist the post-run state so the next run knows what to clean up.
        if only_node is None:
            wf["last_exports"] = {"dir": target_dir.name, "files": sorted(produced)}
            storage.save_workflow(pid, wf)
        _persist_run_record(pid, wid, run_record, run_started_at)
        yield {"event": "done"}
    finally:
        con.close()


def _persist_run_record(pid: str, wid: str, record: dict, started_at: float) -> None:
    """Tag the duration totale + append to the workflow's `runs.jsonl`.
    Best-effort : on n'interrompt jamais un run pour une erreur d'historique."""
    import time as _time

    record["duration_ms"] = int((_time.time() - started_at) * 1000)
    try:
        storage.append_run(pid, wid, record)
    except Exception:  # noqa: BLE001
        pass


def run_workflow(pid, wid, only_node=None) -> dict:
    ran, results = [], {}
    for ev in iter_run_workflow(pid, wid, only_node):
        if ev["event"] == "node_done":
            ran.append(ev["node_id"])
            results[ev["node_id"]] = ev["result"]
        elif ev["event"] == "error":
            raise RuntimeError(ev.get("message", "erreur d'exécution"))
    return {"ran": ran, "results": results}


# --------------------------------------------------------------------------- #
# Preview & profile
# --------------------------------------------------------------------------- #
def _slike(v) -> str:
    return "'%" + str(v).replace("'", "''") + "%'"


def _build_where(columns, filters, q) -> str:
    names = {c["name"] for c in columns}
    clauses = []
    for f in filters or []:
        col = f.get("column")
        if col not in names:
            continue
        op = f.get("op", "contains")
        val = f.get("value")
        ident = q_ident(col)
        if op == "is_null":
            clauses.append(f"{ident} IS NULL")
        elif op == "equals":  # exact match (top-value click)
            if isinstance(val, bool):
                clauses.append(f"{ident} = {'TRUE' if val else 'FALSE'}")
            elif isinstance(val, (int, float)):
                clauses.append(f"{ident} = {val}")
            elif val not in (None, ""):
                clauses.append(
                    f"CAST({ident} AS VARCHAR) = '{str(val).replace(chr(39), chr(39) * 2)}'"
                )
        elif val not in (None, ""):  # contains (text filter input)
            clauses.append(f"CAST({ident} AS VARCHAR) ILIKE {_slike(val)}")
    if q:
        ors = [f"CAST({q_ident(c)} AS VARCHAR) ILIKE {_slike(q)}" for c in names]
        if ors:
            clauses.append("(" + " OR ".join(ors) + ")")
    return " AND ".join(clauses)


def _passthrough_source(pid, wid, node_id, handle):
    """Where to read a node's preview from. An Export block is a pass-through sink
    (it writes a file, transforms nothing) and stores no output of its own — so its
    preview is read straight from its upstream input. Returns (node_id, handle) of
    the parquet to read, falling back to the node itself when no redirect applies."""
    path = storage.node_parquet(pid, wid, node_id, handle)
    if path.exists():
        return node_id, handle
    wf = storage.get_workflow(pid, wid)
    if not wf:
        return node_id, handle
    node = next((n for n in wf.get("nodes", []) if n["id"] == node_id), None)
    if not node or node.get("type") != "export":
        return node_id, handle
    for e in wf.get("edges", []):  # the export's single input edge
        if e.get("target") == node_id:
            return e.get("source"), e.get("sourceHandle") or "out"
    return node_id, handle


def preview_node(
    pid,
    wid,
    node_id,
    handle="out",
    limit=200,
    offset=0,
    sort=None,
    direction="asc",
    filters=None,
    q=None,
) -> dict:
    node_id, handle = _passthrough_source(pid, wid, node_id, handle)
    path = storage.node_parquet(pid, wid, node_id, handle)
    if not path.exists():
        return {"available": False}
    n = _empty_parquet_rows(path)
    if n is not None:  # 0-column 'empty' output: nothing to show
        meta = storage.read_node_meta(pid, wid, node_id, handle) or {}
        return {
            "available": True,
            "columns": [],
            "rows": [],
            "row_count": n,
            "total_count": n,
            "compiled_sql": meta.get("compiled_sql"),
            "clean_report": meta.get("clean_report"),
            "report": meta.get("report"),
            "stats": meta.get("stats"),
            "input_rows": meta.get("input_rows"),
            "unrouted_rows": meta.get("unrouted_rows"),
            "casts_report": meta.get("casts_report"),
            "mixed_types": meta.get("mixed_types"),
        }
    con = duckdb.connect()
    try:
        rel = f"read_parquet({_lit(path)})"
        columns = _describe(con, rel)
        total = con.execute(f"SELECT count(*) FROM {rel}").fetchone()[0]
        where = _build_where(columns, filters, q)
        base = f"SELECT * FROM {rel}"
        if where:
            base += " WHERE " + where
            row_count = con.execute(f"SELECT count(*) FROM ({base})").fetchone()[0]
        else:
            row_count = total
        sql = base
        if sort and any(c["name"] == sort for c in columns):
            d = "DESC" if (direction or "asc").lower() == "desc" else "ASC"
            sql += f" ORDER BY {q_ident(sort)} {d}"
        sql += f" LIMIT {int(limit)} OFFSET {int(offset)}"
        rows = _df_records(con.execute(sql).df())
        meta = storage.read_node_meta(pid, wid, node_id, handle) or {}
        return {
            "available": True,
            "columns": columns,
            "rows": rows,
            "row_count": int(row_count),
            "total_count": int(total),
            "compiled_sql": meta.get("compiled_sql"),
            "clean_report": meta.get("clean_report"),
            "report": meta.get("report"),
            "stats": meta.get("stats"),
            "input_rows": meta.get("input_rows"),
            "unrouted_rows": meta.get("unrouted_rows"),
            "casts_report": meta.get("casts_report"),
            "mixed_types": meta.get("mixed_types"),
        }
    finally:
        con.close()


def column_profile(pid, wid, node_id, column, handle="out") -> dict:
    node_id, handle = _passthrough_source(pid, wid, node_id, handle)
    path = storage.node_parquet(pid, wid, node_id, handle)
    if not path.exists():
        return {"available": False}
    if _empty_parquet_rows(path) is not None:  # 0-column output: no column to profile
        return {"available": False}
    con = duckdb.connect()
    try:
        rel = f"read_parquet({_lit(path)})"
        columns = _describe(con, rel)
        match = next((c for c in columns if c["name"] == column), None)
        if not match:
            return {"available": False}
        ident = q_ident(column)
        total = con.execute(f"SELECT count(*) FROM {rel}").fetchone()[0]
        nulls = con.execute(f"SELECT count(*) FROM {rel} WHERE {ident} IS NULL").fetchone()[0]
        distinct = con.execute(f"SELECT count(DISTINCT {ident}) FROM {rel}").fetchone()[0]
        numeric = _is_numeric_type(match["type"])
        null_pct = round(nulls / total * 100, 1) if total else 0.0
        res = {
            "available": True,
            "type": match["type"],
            "total": int(total),
            "nulls": int(nulls),
            "null_pct": null_pct,
            "distinct": int(distinct),
            "numeric": numeric,
            "numeric_stats": None,
            "histogram": [],
            "top_values": [],
        }

        if numeric and (total - nulls) > 0:
            sub = f"(SELECT {ident} AS c FROM {rel} WHERE {ident} IS NOT NULL)"
            mn, mx, avg, med, sd = con.execute(
                f"SELECT min(c), max(c), avg(c), median(c), stddev_samp(c) FROM {sub}"
            ).fetchone()
            res["numeric_stats"] = {
                "min": _json_safe(mn),
                "max": _json_safe(mx),
                "avg": _json_safe(avg),
                "median": _json_safe(med),
                "stddev": _json_safe(sd),
            }
            mnf, mxf = float(mn), float(mx)
            if mxf > mnf:
                width = (mxf - mnf) / 10
                rows = con.execute(
                    f"SELECT LEAST(9, CAST(floor((c - {mnf}) / {width}) AS INTEGER)) AS b, "
                    f"count(*) FROM {sub} GROUP BY b"
                ).fetchall()
                counts = {int(r[0]): int(r[1]) for r in rows}
                res["histogram"] = [
                    {"lo": mnf + i * width, "hi": mnf + (i + 1) * width, "count": counts.get(i, 0)}
                    for i in range(10)
                ]
            else:
                res["histogram"] = [{"lo": mnf, "hi": mxf, "count": int(total - nulls)}]

        top = con.execute(
            f"SELECT {ident} AS v, count(*) AS c FROM {rel} GROUP BY v ORDER BY c DESC LIMIT 10"
        ).fetchall()
        res["top_values"] = [{"value": _json_safe(r[0]), "count": int(r[1])} for r in top]
        return res
    finally:
        con.close()


def keys_group(pid, wid, node_id, key_cols, q, handle="out", limit=200) -> dict:
    """Drill-down for the « Clés multiples » analysis: returns the rows of the
    materialized node output whose key (any of the key columns, cast to text)
    contains `q`, ordered by the key so the frontend can group them by tuple."""
    node_id, handle = _passthrough_source(pid, wid, node_id, handle)
    path = storage.node_parquet(pid, wid, node_id, handle)
    if not path.exists():
        return {"available": False}
    if _empty_parquet_rows(path) is not None:  # 0-column output: no keys to explore
        return {"available": False}
    key_cols = [c for c in (key_cols or []) if c]
    con = duckdb.connect()
    try:
        rel = f"read_parquet({_lit(path)})"
        columns = _describe(con, rel)
        valid = [c for c in key_cols if any(col["name"] == c for col in columns)]
        if not valid:
            return {"available": False}
        sql = f"SELECT * FROM {rel}"
        text = (q or "").strip()
        if text:
            pat = "'%" + text.replace("'", "''").replace("\\", "\\\\") + "%'"
            ors = " OR ".join(f"CAST({q_ident(c)} AS VARCHAR) ILIKE {pat}" for c in valid)
            sql += f" WHERE {ors}"
        row_count = con.execute(f"SELECT count(*) FROM ({sql})").fetchone()[0]
        sql += " ORDER BY " + ", ".join(q_ident(c) for c in valid)
        sql += f" LIMIT {int(limit)}"
        rows = _df_records(con.execute(sql).df())
        return {
            "available": True,
            "columns": columns,
            "key": valid,
            "rows": rows,
            "row_count": int(row_count),
        }
    finally:
        con.close()
