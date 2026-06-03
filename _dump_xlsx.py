import openpyxl, sys, json
path = r"C:\Users\m.lecouls\Roade\Plan de travail Traitement fichiers.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print("SHEETS:", wb.sheetnames)
for ws in wb.worksheets:
    print("\n" + "="*80)
    print(f"SHEET: {ws.title}  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}")
    print("="*80)
    for row in ws.iter_rows(values_only=True):
        # skip fully empty rows
        if all(c is None or (isinstance(c, str) and c.strip()=="") for c in row):
            continue
        cells = ["" if c is None else str(c) for c in row]
        # trim trailing empties
        while cells and cells[-1]=="":
            cells.pop()
        print(" | ".join(cells))
