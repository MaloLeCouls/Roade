"""Tests de non-régression « intégrité FR » (todo A.11).

Ces tests sont la digue contre le bug le plus dangereux de Roade : un CSV FR mal
lu (encodage, virgule décimale) rend des chiffres faux *sans rien dire*. Cf.
audit 08 risque #1 et la doctrine anti-slop : le slop produit du plausible-mais-faux ;
ici on garantit que les nombres sortent justes ou qu'on s'en aperçoit.

Cas couverts :
    - encodage CP1252 (Excel FR) avec accents : pas de mojibake
    - virgule décimale FR : `12,5 + 7,3 = 19.8` (et non `0`)
    - séparateur `;` reconnu automatiquement
    - fingerprint change quand les options de lecture changent (A.9)
"""

from __future__ import annotations

import pandas as pd

import engine
import storage


def _new_project(name: str) -> tuple[str, object]:
    p = storage.create_project(name)
    return p["id"], storage.files_dir(p["id"])


def test_cp1252_no_mojibake(tmp_path):
    """Un CSV encodé CP1252 avec accents (« Référence », « Montant ») s'importe
    avec les accents intacts. Sans sniffing, pandas (utf-8 par défaut) crasherait
    ou renverrait `R\\xe9f\\xe9rence` côté contenu."""
    pid, fd = _new_project("CP1252")
    csv_text = "Référence;Pays\nABC;Bénin\nXYZ;Côte d'Ivoire\n"
    (fd / "in.csv").write_bytes(csv_text.encode("cp1252"))

    info = engine.peek_source(pid, "in.csv", None, 0)
    assert info["detected"]["encoding"].lower().replace("-", "").startswith("cp1252") or (
        info["detected"]["encoding"].lower() in ("windows-1252", "iso-8859-1")
    ), info["detected"]
    cols = [c["name"] for c in info["columns"]]
    assert "Référence" in cols, cols  # accent préservé dans l'en-tête


def test_fr_decimal_sum_is_correct(tmp_path):
    """`Article;Prix\\nA;12,5\\nB;7,3` → SUM = 19.8 (pas 0, pas string).
    C'est le scénario littéral de l'audit 08 § 2.2 et le test régression #1."""
    pid, fd = _new_project("FRdec")
    (fd / "p.csv").write_text("Article;Prix\nA;12,5\nB;7,3\n", encoding="utf-8")

    # Sniffing transparent : pas de data utilisateur, on se repose sur l'auto.
    info = engine.peek_source(pid, "p.csv", None, 0)
    assert info["detected"].get("decimal") == ",", info["detected"]

    # Workflow : Source -> SQL (SUM(Prix))
    wf = storage.create_workflow(pid, "W")
    wid = wf["id"]
    wf["nodes"] = [
        {"id": "s", "type": "source", "position": {}, "data": {"file": "p.csv"}},
        {
            "id": "agg",
            "type": "sql",
            "position": {},
            "data": {
                "mode": "raw",
                "raw_sql": 'SELECT SUM("Prix") AS total FROM in1',
            },
        },
    ]
    wf["edges"] = [
        {"id": "e1", "source": "s", "target": "agg", "sourceHandle": "out", "targetHandle": "in1"}
    ]
    storage.save_workflow(pid, wf)
    engine.run_workflow(pid, wid)

    rows = engine.preview_node(pid, wid, "agg")["rows"]
    assert rows and float(rows[0]["total"]) == 19.8, rows


def test_separator_detected_in_peek(tmp_path):
    """L'aperçu remonte le séparateur `;` dans le `detected` (A.3)."""
    pid, fd = _new_project("Sep")
    (fd / "f.csv").write_text("a;b;c\n1;2;3\n4;5;6\n", encoding="utf-8")
    info = engine.peek_source(pid, "f.csv", None, 0)
    assert info["detected"].get("sep") == ";", info["detected"]


def test_user_override_beats_sniffer(tmp_path):
    """Si l'utilisateur force `decimal=.`, on ne re-sniffe pas — la valeur passe
    telle quelle à pandas (preuve qu'on ne joue pas le devin par-dessus)."""
    pid, fd = _new_project("Override")
    (fd / "p.csv").write_text("Article;Prix\nA;12,5\n", encoding="utf-8")

    info = engine.peek_source(pid, "p.csv", None, 0, data={"decimal": "."})
    # decimal n'est PAS dans detected (l'utilisateur a choisi), mais aucun crash
    assert "decimal" not in info["detected"]


def test_csv_export_excel_fr_preset_roundtrip(tmp_path):
    """Le preset Excel FR (sep=';' + decimal=',' + BOM) doit produire un CSV
    que pandas relit avec accents intacts et nombres typés (cf. todo A.4).
    Le BOM est ce qui fait qu'Excel FR autodétecte UTF-8 sans demander."""
    df = pd.DataFrame({"Référence": ["Côte"], "Prix": [12.5]})
    out = tmp_path / "out.csv"
    engine._write_csv(df, out, {})

    raw = out.read_bytes()
    assert raw[:3] == b"\xef\xbb\xbf", "BOM UTF-8 attendu pour Excel FR"
    assert b";" in raw and b"12,5" in raw

    back = pd.read_csv(out, sep=";", decimal=",", encoding="utf-8-sig")
    assert list(back.columns) == ["Référence", "Prix"]
    assert back.iloc[0]["Prix"] == 12.5
    assert back.iloc[0]["Référence"] == "Côte"


def test_csv_export_respects_overrides(tmp_path):
    """L'utilisateur peut renverser le preset (sep `,`, decimal `.`, no BOM)."""
    df = pd.DataFrame({"a": [1.5], "b": [2.5]})
    out = tmp_path / "en.csv"
    engine._write_csv(df, out, {"sep": ",", "decimal": ".", "bom": False, "eol": "\n"})
    raw = out.read_bytes()
    assert not raw.startswith(b"\xef\xbb\xbf")
    assert b"," in raw and b"1.5" in raw and b"\r\n" not in raw


def test_excel_export_refuses_beyond_limits():
    """A.7 — un export Excel hors limites doit échouer AVEC un message
    actionnable, AVANT toute écriture (pas après une heure de calcul)."""
    import pytest as _pytest

    # Trop de colonnes : on construit le shape sans matérialiser les valeurs
    df_wide = pd.DataFrame({f"c{i}": [0] for i in range(engine.EXCEL_MAX_COLS + 5)})
    with _pytest.raises(ValueError, match="colonnes"):
        engine._check_excel_limits(df_wide)

    # Trop de lignes — on simule sans allouer 1M lignes
    class _FakeDF:
        columns = ["a"]

        def __len__(self):
            return engine.EXCEL_MAX_ROWS + 1

    with _pytest.raises(ValueError, match="lignes"):
        engine._check_excel_limits(_FakeDF())


def test_cast_to_number_with_fr_comma(tmp_path):
    """A.5 — l'utilisateur peut forcer une colonne en `number` sans bloc Nettoyage.
    Les ratés sont comptés (`failed`) et exposés dans le résultat du bloc."""
    pid, fd = _new_project("CastNum")
    (fd / "p.csv").write_text("Article;Prix\nA;12,5\nB;7,3\nC;abc\n", encoding="utf-8")

    wf = storage.create_workflow(pid, "W")
    wid = wf["id"]
    wf["nodes"] = [
        {
            "id": "s",
            "type": "source",
            "position": {},
            "data": {
                "file": "p.csv",
                # decimal sniffé en `,` (cf. A.2) ; on force juste un cast number.
                "casts": [{"column": "Prix", "type": "number"}],
            },
        }
    ]
    wf["edges"] = []
    storage.save_workflow(pid, wf)
    res = engine.run_workflow(pid, wid)
    rep = res["results"]["s"].get("casts_report")
    assert rep == [{"column": "Prix", "type": "number", "failed": 1}], rep


def test_cast_to_date_fr(tmp_path):
    """A.6 — une colonne `JJ/MM/AAAA` typée en date est triable chronologiquement."""
    pid, fd = _new_project("CastDate")
    (fd / "p.csv").write_text(
        "Doc;Date\nA;05/03/2024\nB;12/01/2024\nC;01/12/2023\n", encoding="utf-8"
    )

    wf = storage.create_workflow(pid, "W")
    wid = wf["id"]
    wf["nodes"] = [
        {
            "id": "s",
            "type": "source",
            "position": {},
            "data": {
                "file": "p.csv",
                "casts": [{"column": "Date", "type": "date", "format": "%d/%m/%Y"}],
            },
        }
    ]
    wf["edges"] = []
    storage.save_workflow(pid, wf)
    engine.run_workflow(pid, wid)

    # Tri chronologique : la plus ancienne (01/12/2023) doit sortir en tête.
    srt = engine.preview_node(pid, wid, "s", sort="Date", direction="asc")
    docs = [r["Doc"] for r in srt["rows"]]
    assert docs == ["C", "B", "A"], docs


def test_cast_boolean_fr_aliases(tmp_path):
    """A.5 — les valeurs FR (`oui`/`non`/`vrai`/`faux`) sont reconnues."""
    pid, fd = _new_project("CastBool")
    (fd / "p.csv").write_text(
        "Article;Actif\nA;oui\nB;non\nC;VRAI\nD;peut-être\n", encoding="utf-8"
    )
    wf = storage.create_workflow(pid, "W")
    wid = wf["id"]
    wf["nodes"] = [
        {
            "id": "s",
            "type": "source",
            "position": {},
            "data": {
                "file": "p.csv",
                "casts": [{"column": "Actif", "type": "boolean"}],
            },
        }
    ]
    wf["edges"] = []
    storage.save_workflow(pid, wf)
    res = engine.run_workflow(pid, wid)
    rep = res["results"]["s"]["casts_report"][0]
    assert rep["column"] == "Actif" and rep["failed"] == 1, rep  # "peut-être"


def test_xlsx_merged_cells_warning(tmp_path):
    """A.8 — un xlsx avec cellules fusionnées dans l'en-tête doit générer un
    avertissement clair, pas un résultat trompeur silencieux."""
    import openpyxl

    pid, fd = _new_project("XlsxMerged")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, "Catégorie A")
    ws.merge_cells("A1:C1")  # fusion ligne 1 colonnes A-C
    ws.append(["nom", "prix", "stock"])
    ws.append(["a", 1, 10])
    wb.save(fd / "x.xlsx")
    wb.close()

    info = engine.peek_source(pid, "x.xlsx", None, 0)
    warns = info.get("warnings") or []
    assert any("fusionn" in w.lower() for w in warns), warns


def test_xlsx_mixed_text_and_numbers_imports_without_crash(tmp_path):
    """Un xlsx « humain » mélange souvent texte et nombres dans la même colonne
    (ex. un libellé qui contient parfois juste `7`). openpyxl renvoie les types
    natifs ; pyarrow refuse alors la colonne object hétérogène (« Expected
    bytes, got a 'int' object ») et `to_parquet` plante.

    Le moteur doit détecter ces colonnes mixtes, les coercer en texte (en
    préservant les vides), surfacer un rapport, et permettre la matérialisation
    parquet. Régression : un Excel client réel a planté avec ce message exact."""
    import openpyxl

    pid, fd = _new_project("MixedTypes")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["code", "libelle"])
    ws.append(["A1", "fermeuse"])
    ws.append(["A2", "ressort"])
    ws.append(["A3", 7])  # nombre orphelin au milieu de libellés texte
    ws.append(["A4", "carter"])
    ws.append(["A5", None])  # vide
    ws.append(["A6", 12])  # un autre nombre
    wb.save(fd / "mixed.xlsx")
    wb.close()

    node = {"id": "s", "type": "source", "data": {"file": "mixed.xlsx"}}
    outputs, extra = engine._run_source(None, pid, node, [])
    df = outputs["out"]

    # La colonne hétérogène a été coercée en texte ; les nombres orphelins sont
    # devenus "7" et "12" plutôt que de faire planter pyarrow. Les vides
    # restent vides (None ou NaN — pandas peut osciller, parquet écrit null
    # dans les deux cas, c'est ce qui compte).
    vals = df["libelle"].tolist()
    expected_non_null = {0: "fermeuse", 1: "ressort", 2: "7", 3: "carter", 5: "12"}
    for i, v in expected_non_null.items():
        assert vals[i] == v, (i, vals[i])
    assert pd.isna(vals[4]) or vals[4] is None, vals[4]

    # Rapport remonté pour que l'UI puisse prévenir l'utilisateur.
    assert "mixed_types" in extra
    rep = {it["column"]: it for it in extra["mixed_types"]}
    assert "libelle" in rep
    assert rep["libelle"]["counts"].get("texte") == 3
    assert rep["libelle"]["counts"].get("nombre") == 2

    # Le parquet s'écrit maintenant sans lever d'exception.
    import tempfile

    p = tmp_path / "out.parquet"
    df.to_parquet(p, index=False)
    assert p.exists() and p.stat().st_size > 0


def test_fingerprint_invalidated_by_read_options(tmp_path):
    """Changer encoding ou decimal change la fingerprint Source → cache invalidé
    (A.9). Le bug serait un cache hit avec un décodage différent."""
    pid, fd = _new_project("Fp")
    (fd / "x.csv").write_text("a;b\n1;2\n", encoding="utf-8")
    node_a = {"id": "s", "type": "source", "data": {"file": "x.csv"}}
    node_b = {"id": "s", "type": "source", "data": {"file": "x.csv", "encoding": "cp1252"}}
    node_c = {"id": "s", "type": "source", "data": {"file": "x.csv", "decimal": ","}}
    fp_a = engine._source_fingerprint(pid, node_a)
    fp_b = engine._source_fingerprint(pid, node_b)
    fp_c = engine._source_fingerprint(pid, node_c)
    assert fp_a and fp_b and fp_c
    assert fp_a != fp_b, "changer encoding doit changer la fingerprint"
    assert fp_a != fp_c, "changer decimal doit changer la fingerprint"
    assert fp_b != fp_c
